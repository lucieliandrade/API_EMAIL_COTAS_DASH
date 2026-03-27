
import pandas as pd
from datetime import datetime
import pyodbc
import numpy as np
import win32com.client as win32
# import requests
import openpyxl
# from bs4 import BeautifulSoup
import tkinter as tk 
import subprocess
import os
from pathlib import Path
from msoffice2pdf import convert
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException



diretorio = "Z:\\Relações com Investidores - NOVO\\codigos\\cotas"


## modelos de mailer
################ Modelos Mailer ##################


def get_fundos():

    df = pd.read_excel("X:\\BDM\\Novo Modelo de Carteiras\\Tipo_Fundos.xlsx", usecols='A,E,F,I,J,K,L')

    df = df.sort_values(by = ['fundo'])

    return df.loc[df['Encerrado'].isna(), ['fundo', 'ADM', 'modelo_mailer', 'cota_inicial_data', 'cota_inicial_valor', 'bench']]

fundo_infos = get_fundos()

fundo_infos.loc[fundo_infos['fundo']=='CAPITANIA FCOPEL', 'fundo'] = 'FCopel'


################ Modelo 1 ################## = Retorno  |  benchmark  |  % benchmark
modelo_1 = fundo_infos.loc[fundo_infos['modelo_mailer']==1, 'fundo'].tolist()

################ Modelo 2 ################## =  Retorno | IPCA + (i) | CDI | % CDI
modelo_2 = fundo_infos.loc[fundo_infos['modelo_mailer']==2, 'fundo'].tolist()

################ Modelo 3 ################## =  Retorno | IFIX
modelo_3 = fundo_infos.loc[fundo_infos['modelo_mailer']==3, 'fundo'].tolist()

################ Modelo 4 ################## = Retorno | CDI | CDI+(i)| % CDI | % CDI+(i)
modelo_4 = fundo_infos.loc[fundo_infos['modelo_mailer']==4, 'fundo'].tolist()

################ Modelo 5 ################## = Retorno | CDI | IMAB
modelo_5 = fundo_infos.loc[fundo_infos['modelo_mailer']==5, 'fundo'].tolist()

################ Modelo Fapes ################## = Retorno | IMAB | Spread IMAB |  IMAB 5  | Spread IMAB 5
modelo_fapes = fundo_infos.loc[fundo_infos['modelo_mailer']=='fapes', 'fundo'].tolist()

################ Modelo FUNCEF ################## =  Retorno | IPCA+7.48% | % IPCA+7.48% | IFIX
modelo_funcef = fundo_infos.loc[fundo_infos['modelo_mailer']=='funcef', 'fundo'].tolist()


################################## Fundos por ADM ##################################

fundo_infos.dropna(subset='modelo_mailer', inplace=True)

fundos_mellon = fundo_infos.loc[fundo_infos['ADM']=='BNYM', 'fundo'].tolist()

fundos_bradesco = fundo_infos.loc[fundo_infos['ADM']=='Bradesco', 'fundo'].tolist()

fundos_itau = fundo_infos.loc[fundo_infos['ADM']=='Itau', 'fundo'].tolist()

fundos_xp = fundo_infos.loc[fundo_infos['ADM']=='XP', 'fundo'].tolist()

fundos_btg = fundo_infos.loc[fundo_infos['ADM']=='BTG', 'fundo'].tolist()

tudo = fundos_mellon + fundos_bradesco + fundos_itau + fundos_xp + fundos_btg


################################## Fundos que tiveram troca de benchmark ##################################
mais_de_1_bench = ['CAPIT PREVI GM', 'CapitâniaQP4']

################################## Fundos com amortização (cota ajustada) ##################################
fundos_com_ajuste = ['XP INFRA90', 'CAPITANIA CORP FIDC']


################################## DATAS ##################################

# Cota Inicial 
def infos_cota(df_):
    
    # Inicializa o dicionário vazio
    result_dict = {}
    
    # Itera sobre as linhas do DataFrame
    for _, row in df_.iterrows():
        chave = row['fundo']
        data = row['cota_inicial_data']
        valor = row['cota_inicial_valor']
       
        # Trata o caso em que valor pode ser None/NaN

        result_dict[chave] = [data, float(valor)]
    
    return result_dict

cota_inicial  = infos_cota(fundo_infos)


################################## benchmarks por fundo ##################################

# Cota Inicial 
def infos_bench(df_):
    
    # Inicializa o dicionário vazio
    result_dict = {}
    
    # Itera sobre as linhas do DataFrame
    for _, row in df_.iterrows():
        chave = row['fundo']
        valor = row['bench']
       
        # Trata o caso em que valor pode ser None/NaN
        if chave in mais_de_1_bench:

            lista_valor = valor.split('$')

            lista_valor_ajustado = [float(i) if i.replace('.', '', 1).isdigit() else i for i in lista_valor]

            result_dict[chave] = tuple(lista_valor_ajustado)

        else:

            result_dict[chave] = valor
    
    return result_dict

fundo_bench  = infos_bench(fundo_infos)


# ano, mês e dia de D-1 

hoje = datetime.today().strftime(format='%Y-%m-%d')
dias_uteis = pd.read_excel('X:\\Leonardo Salles\\Gestão\Relatorio\\D_Uteis.xlsx')
dmenos1 = dias_uteis[dias_uteis['Dias Uteis'] < hoje]['Dias Uteis'].iloc[-1]

ano = dmenos1[:4]
mes = dmenos1[5:7]
dia = dmenos1[-2:]

# retorna D-X
def dmenos(dus):

    hoje = datetime.today().strftime(format='%Y-%m-%d')
    
    return dias_uteis[dias_uteis['Dias Uteis'] < hoje]['Dias Uteis'].iloc[-dus]

# fechamento
d_corte = datetime.strptime(dmenos1,'%Y-%m-%d').replace(day=1).strftime(format='%Y-%m-%d')
fechamento = dias_uteis[dias_uteis['Dias Uteis'] < d_corte]['Dias Uteis'].iloc[-1]
# outras datas
ano_novo = datetime.strptime(dmenos1,'%Y-%m-%d').replace(day=1,month=1).strftime(format='%Y-%m-%d')
ano_velho = dias_uteis[dias_uteis['Dias Uteis'] < ano_novo]['Dias Uteis'].iloc[-1]
end_ano_menos2 = dias_uteis[dias_uteis['Dias Uteis'].str[:4]==str(int(dmenos1[:4])-2)].iloc[-1,1]
end_ano_menos3 = dias_uteis[dias_uteis['Dias Uteis'].str[:4]==str(int(dmenos1[:4])-3)].iloc[-1,1]
end_ano_menos4 = dias_uteis[dias_uteis['Dias Uteis'].str[:4]==str(int(dmenos1[:4])-4)].iloc[-1,1]




def remove_trailing_zeros(num_str: str) -> str:
    """
    Remove zeros desnecessários à direita de um número representado como string.
    
    :param num_str: Número em formato de string.
    :return: Número sem zeros desnecessários à direita.
    """
    if '.' in num_str:
        return num_str.rstrip('0').rstrip('.')  # Remove zeros e ponto decimal, se necessário
    return num_str  # Retorna diretamente se não houver ponto decimal



################################## Dados Carteira ##################################
def cota_carteira(fundo_):
    
    cart_itau_sem_cdi = ['CAPITANIA CANA', 'FCopel_Imob', 'Sabesprev']

    if fundo_ in fundos_mellon:
    
        # Carrega a planilha em um objeto DataFrame do pandas
        planilha = pd.read_excel(f"X:\\#CapitaniaRFE\\Operational\\BatimentoCotas\\Carteiras_BNYM\\{fundo_}_{ano}{mes}{dia}.xlsx")

        # Encontra a linha que contém a Célula X com valor Y
        
        linha_cota = planilha.index[planilha['Cota Liberada'] == 'Valor da Cota Liquida'][0]
        
        linha_dia = planilha.index[planilha['Tipo da Posição']=='Variação Diária'][0]
        
        linha_mes = planilha.index[planilha['Moeda']=='Variação Mensal'][0]
        
        linha_ano = planilha.index[planilha['Cotação']=='Variação Anual'][0]
        
        
        #Cota
        valor_cota = planilha.at[linha_cota + 1, 'Cota Liberada']
        
        #Dia
        var_dia = planilha.at[linha_dia + 1, 'Tipo da Posição']
        dia_cdi = planilha.loc[planilha['Cliente']=='CDI' , 'Tipo da Posição'].values[0]
        p_cdi = planilha.loc[planilha['Cliente']=='CDI' , 'Data da Posição'].values[0]
        
        #Mês
        var_mes = planilha.at[linha_mes + 1, 'Moeda']
        mes_cdi = planilha.loc[planilha['Cliente']=='CDI' , 'Moeda'].values[0]
        
        #Ano
        var_ano = planilha.at[linha_ano + 1, 'Cotação']
        ano_cdi = planilha.loc[planilha['Cliente']=='CDI' , 'Cotação'].values[0]

        #Criando Dicionário Resultado
        dados_carteira = {}
        
        dados_carteira['valor_cota'] = valor_cota
        
        
        dados_carteira['var_dia'] = var_dia
        dados_carteira['p_cdi'] = p_cdi
        dados_carteira['dia_cdi'] = dia_cdi
            
        dados_carteira['var_mes'] = var_mes
        dados_carteira['mes_cdi'] = mes_cdi
        
        dados_carteira['var_ano'] = var_ano
        dados_carteira['ano_cdi'] = ano_cdi
        
        return dados_carteira


    elif fundo_ in fundos_btg:

        # Carrega a planilha em um objeto DataFrame do pandas
        planilha = pd.read_excel(f"X:\\#CapitaniaRFE\\Operational\\BatimentoCotas\\Carteiras_BTGPactual\\{fundo_}_{dia}_{mes}_{ano}.xlsx", skiprows= 15)

        # Encontra a linha que contém a Célula X com valor Y
        i = planilha.loc[planilha[planilha.columns[1]]=='Cota Líquida', planilha.columns[1]].index[0]

        valor_cota = planilha.iloc[i+1, 1]
        
        var_dia = planilha.iloc[i+1, 2] / 100
        p_cdi_dia = planilha.iloc[i+1, 6]   / 100
            
        var_mes = planilha.iloc[i+1, 3]  / 100
        p_cdi_mes = planilha.iloc[i+1, 7] / 100
        
        var_ano = planilha.iloc[i+1, 5]  / 100
        p_cdi_ano = planilha.iloc[i+1, 9] / 100

        #Criando Dicionário Resultado
        dados_carteira = {}
        
        dados_carteira['valor_cota'] = valor_cota
        
        
        dados_carteira['var_dia'] = var_dia
        dados_carteira['p_cdi'] = p_cdi_dia
                    
        dados_carteira['var_mes'] = var_mes
        dados_carteira['p_cdi_mes'] = p_cdi_mes
        
        dados_carteira['var_ano'] = var_ano
        dados_carteira['p_cdi_ano'] = p_cdi_ano
        
        return dados_carteira
    
    
    elif fundo_ in fundos_itau and fundo_ not in cart_itau_sem_cdi:

        # Carrega a planilha em um objeto DataFrame do pandas
        planilha1 = pd.read_excel(f"X:\#CapitaniaRFE\Operational\BatimentoCotas\Carteiras_Intrag\{fundo_}_PosicaoDiaria{ano}{mes}{dia}.xlsx", sheet_name='Patrimonio_Cotas', nrows=2, usecols='A:M')
        planilha2 = pd.read_excel(f"X:\#CapitaniaRFE\Operational\BatimentoCotas\Carteiras_Intrag\{fundo_}_PosicaoDiaria{ano}{mes}{dia}.xlsx", sheet_name='Rentabilidade', nrows=3, usecols='C:G')


        valor_cota = planilha1.iloc[0,12]

        var_dia = planilha2.iloc[0,2] / 100
        dia_cdi = planilha2.iloc[1,2] / 100
        p_cdi = planilha2.iloc[1,0] / 100
            
        var_mes = planilha2.iloc[0, 3]  / 100
        mes_cdi = planilha2.iloc[1,3] / 100

        var_ano = planilha2.iloc[0, 4]  / 100
        ano_cdi = planilha2.iloc[1,4] / 100

        #Criando Dicionário Resultado
        dados_carteira = {}
        
        dados_carteira['valor_cota'] = valor_cota
        
        
        dados_carteira['var_dia'] = var_dia
        dados_carteira['p_cdi'] = p_cdi
        dados_carteira['dia_cdi'] = dia_cdi
            
        dados_carteira['var_mes'] = var_mes
        dados_carteira['mes_cdi'] = mes_cdi
        
        dados_carteira['var_ano'] = var_ano
        dados_carteira['ano_cdi'] = ano_cdi
        
        return dados_carteira

    elif fundo_ in cart_itau_sem_cdi:

        # Carrega a planilha em um objeto DataFrame do pandas
        planilha1 = pd.read_excel(f"X:\#CapitaniaRFE\Operational\BatimentoCotas\Carteiras_Intrag\{fundo_}_PosicaoDiaria{ano}{mes}{dia}.xlsx", sheet_name='Patrimonio_Cotas', nrows=2, usecols='A:M')
        planilha2 = pd.read_excel(f"X:\#CapitaniaRFE\Operational\BatimentoCotas\Carteiras_Intrag\{fundo_}_PosicaoDiaria{ano}{mes}{dia}.xlsx", sheet_name='Rentabilidade', nrows=3, usecols='C:G')


        valor_cota = planilha1.iloc[0,12]

        var_dia = planilha2.iloc[0,2] / 100
            
        var_mes = planilha2.iloc[0, 3]  / 100
        
        var_ano = planilha2.iloc[0, 4]  / 100
        

        #Criando Dicionário Resultado
        dados_carteira = {}
        
        dados_carteira['valor_cota'] = valor_cota
        
        dados_carteira['var_dia'] = var_dia
            
        dados_carteira['var_mes'] = var_mes
        
        dados_carteira['var_ano'] = var_ano
        
        return dados_carteira
    
    
    elif fundo_ in fundos_bradesco:

        # Carrega a planilha em um objeto DataFrame do pandas
        planilha = pd.read_excel(f"X:\\#CapitaniaRFE\\Operational\\BatimentoCotas\\Carteiras_Bradesco\\{fundo_}_{ano}{mes}{dia}.xls")

        # Encontra a linha que contém a Célula X com valor Y
        
        linha_cota = planilha.index[planilha[planilha.columns[0]] == 'Valor da cota unitária (Líquida)'][0]
        
        linha_r_cota = planilha.index[planilha[planilha.columns[0]]=='COTA'][0]
        
        linha_r_cdi = planilha.index[planilha[planilha.columns[0]]=='CDI'][0]
        
        
        #Cota
        valor_cota = planilha.iloc[linha_cota, 1]
        
        #Dia
        var_dia = planilha.iloc[linha_r_cota, 3] / 100
        dia_cdi = planilha.iloc[linha_r_cdi, 3] / 100
        p_cdi = planilha.iloc[linha_r_cdi, 1] / 100
        
        #Mês
        var_mes = planilha.iloc[linha_r_cota, 4] / 100
        mes_cdi = planilha.iloc[linha_r_cdi, 4] / 100
        
        #Ano
        var_ano = planilha.iloc[linha_r_cota, 5] / 100
        ano_cdi = planilha.iloc[linha_r_cdi, 5] / 100

        #Criando Dicionário Resultado
        dados_carteira = {}
        
        dados_carteira['valor_cota'] = valor_cota
        
        
        dados_carteira['var_dia'] = var_dia
        dados_carteira['p_cdi'] = p_cdi
        dados_carteira['dia_cdi'] = dia_cdi
            
        dados_carteira['var_mes'] = var_mes
        dados_carteira['mes_cdi'] = mes_cdi
        
        dados_carteira['var_ano'] = var_ano
        dados_carteira['ano_cdi'] = ano_cdi
        
        return dados_carteira
    
    elif fundo_ in fundos_xp:
        
        if fundo_ == 'CAPITANIA YIELD 120':
            planilha = pd.read_excel(f"X:\\#CapitaniaRFE\\Operational\\BatimentoCotas\\Carteiras_XP\\Carteira_diária_YIELD_120_{ano}-{mes}-{dia}.xlsx", sheet_name='Rentabilidade (%)')
        else:
            # Carrega a planilha em um objeto DataFrame do pandas
            planilha = pd.read_excel(f"X:\\#CapitaniaRFE\\Operational\\BatimentoCotas\\Carteiras_XP\\Carteira_diária_{fundo_}_{ano}-{mes}-{dia}.xlsx", sheet_name='Rentabilidade (%)')

        valor_cota = float(planilha.iloc[0,3])
        var_dia = float(planilha.iloc[0,0].replace('%', '')) / 100
        var_mes = float(planilha.iloc[0,1].replace('%', '')) / 100
        var_ano = float(planilha.iloc[0,2].replace('%', '')) / 100

        #Criando Dicionário Resultado
        dados_carteira = {}
        
        dados_carteira['valor_cota'] = valor_cota
        dados_carteira['var_dia'] = var_dia
        dados_carteira['var_mes'] = var_mes    
        dados_carteira['var_ano'] = var_ano

            
        return dados_carteira
    
    
################################## Importação dos Índices ################################### 
################################## IMA ##################################

conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};'
                      'SERVER=rds01.capitania.net;'
                      'DATABASE=db_asset;UID=mcavalcanti;PWD=H0OO^08z9^$#;')
query = ('SELECT * FROM db_asset.dbo.ANBIMA_IMA order by data_referencia')

imas = pd.read_sql(sql=query, con=conn)
imas['data_referencia'] = pd.to_datetime(imas['data_referencia'])
imas['data_referencia'] = imas['data_referencia'].dt.strftime('%Y-%m-%d')


new_imab = imas.loc[imas['indice'] == 'IMA-B', ['data_referencia', 'numero_indice']]
new_imab.rename(columns={'data_referencia':'Data', 'numero_indice':'Acum'}, inplace=True)

new_imab5 = imas.loc[imas['indice'] == 'IMA-B 5', ['data_referencia', 'numero_indice']]
new_imab5.rename(columns={'data_referencia':'Data', 'numero_indice':'Acum'}, inplace=True)

old_imab = pd.read_excel("Z:\Relações com Investidores - NOVO\codigos\indices\IMAB-HISTORICO.xls", usecols='B:C')
old_imab['Data de Referência'] = pd.to_datetime(old_imab['Data de Referência'])
old_imab['Data de Referência'] = old_imab['Data de Referência'].dt.strftime('%Y-%m-%d')
old_imab.rename(columns={'Data de Referência':'Data', 'Número Índice':'Acum'}, inplace=True)

old_imab5 = pd.read_excel("Z:\Relações com Investidores - NOVO\codigos\indices\IMAB5-HISTORICO.xls", usecols='B:C')
old_imab5['Data de Referência'] = pd.to_datetime(old_imab5['Data de Referência'])
old_imab5['Data de Referência'] = old_imab5['Data de Referência'].dt.strftime('%Y-%m-%d')
old_imab5.rename(columns={'Data de Referência':'Data', 'Número Índice':'Acum'}, inplace=True)

# IMAB e IMAB5 acumulados
imab_acum = pd.concat([old_imab.loc[old_imab['Data']<new_imab.iloc[0,0]] , new_imab])
imab5_acum = pd.concat([old_imab5.loc[old_imab5['Data']<new_imab5.iloc[0,0]] , new_imab5])



################################## IPCA ##################################

#obtendo o IPCA efetivo 
query = ('SELECT * FROM db_asset.dbo.INDICE_INFLACAO order by data')
ipca_base = pd.read_sql(sql=query, con=conn)
ipca_base['data'] = pd.to_datetime(ipca_base['data'])
ipca_base['data'] = ipca_base['data'].dt.strftime('%Y-%m-%d')
ipca_base = ipca_base.loc[ipca_base['indexador']=='IPCA', ['data', 'inf_mensal']]

dates = dias_uteis[(dias_uteis['Dias Uteis'] < hoje) & (dias_uteis['Dias Uteis'] >= ipca_base.iloc[0,0])][['Dias Uteis']]

ipca_base['mes'] = ipca_base['data'].str[:7]
dates['mes'] = dates['Dias Uteis'].str[:7]

#após a linha abaixo, temos o IPCA para todos os dias úteis,menos para os dias mais recentes, que retornam NaN
ipca = dates.merge(ipca_base[['mes', 'inf_mensal']], how = 'left', on='mes')
################################## Projeção do IPCA ##################################

m0_extenso = {
    1 :'JANEIRO',
    2 :'FEVEREIRO',
    3 :'MARÇO',
    4 :'ABRIL',
    5 :'MAIO',
    6 :'JUNHO',
    7 :'JULHO',
    8 :'AGOSTO',
    9 :'SETEMBRO',
    10 :'OUTUBRO',
    11 :'NOVEMBRO',
    12 :'DEZEMBRO'}
    

# retorna a última projeção (q no início do mês ainda não consta na tabela SQL do Picci)


def web_scraping_anbima():
    """
    Função que realiza web scraping dos dados de projeção de inflação da ANBIMA.
    Extrai dados da coluna 'Projeção (%)' de duas tabelas específicas e retorna
    os dados da tabela apropriada baseado no mês atual.
    
    Lógica condicional:
    - SE o mês atual (m0_extenso[int(mes)]) estiver contido nos cabeçalhos 'th' da primeira tabela,
      ENTÃO retorna os dados da primeira tabela
    - SENÃO retorna os dados da segunda tabela
    
    Returns:
        list: Lista com os dados da coluna "Projeção (%)" da tabela selecionada
              Formato: ['Projeção (%)', valor1, valor2, ...]
    """
    url = "https://www.anbima.com.br/pt_br/informar/estatisticas/precos-e-indices/projecao-de-inflacao-gp-m.htm"
    
    # Configurar opções do Chrome
    chrome_options = Options()
    chrome_options.add_argument("--headless")  # Executar em modo headless
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)
    
    driver = None
    try:
        # Inicializar o driver
        driver = webdriver.Chrome(options=chrome_options)
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        
        # print(f"Carregando URL: {url}")
        driver.get(url)
        
        # Aguardar um pouco para a página carregar completamente
        import time
        import re
        time.sleep(5)
        
        # print(f"Título da página: {driver.title}")
        # print(f"URL atual: {driver.current_url}")
        
        # Aguardar o elemento profile carregar
        wait = WebDriverWait(driver, 30)
        try:
            wait.until(EC.presence_of_element_located((By.ID, "profile")))
            # print("Elemento 'profile' encontrado!")
            
            # Aguardar especificamente pelas tabelas terem conteúdo
            time.sleep(3)
            try:
                wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="profile"]/div/div[1]/table//td[text()]')))
                # print("Conteúdo das tabelas carregado!")
            except TimeoutException:
                # print("Timeout aguardando conteúdo das tabelas, mas continuando...")
                # Tentar forçar o carregamento com JavaScript
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(2)
                driver.execute_script("window.scrollTo(0, 0);")
                time.sleep(1)
        except TimeoutException:
            # print("Elemento 'profile' não encontrado. Tentando localizar outros elementos...")
            # Tentar encontrar qualquer tabela na página
            tables = driver.find_elements(By.TAG_NAME, "table")
            # print(f"Número de tabelas encontradas na página: {len(tables)}")
            
            # Verificar se existe algum elemento com ID similar
            profile_elements = driver.find_elements(By.CSS_SELECTOR, "[id*='profile']")
            # print(f"Elementos com 'profile' no ID: {len(profile_elements)}")
            # for elem in profile_elements:
                # print(f"  - ID: {elem.get_attribute('id')}")
            
            # Se não encontrar o profile, tentar continuar mesmo assim
            if len(tables) == 0:
                raise TimeoutException("Nenhuma tabela encontrada na página")
        
        # XPaths das tabelas
        table1_xpath = '//*[@id="profile"]/div/div[1]/table'
        table2_xpath = '//*[@id="profile"]/div/div[3]/table'
        
        # Função auxiliar para extrair dados da coluna "Projeção (%)"
        def extract_projecao_column(table_xpath):
            try:
                # Aguardar a tabela estar presente
                wait.until(EC.presence_of_element_located((By.XPATH, table_xpath)))
                table = driver.find_element(By.XPATH, table_xpath)
                
                # print(f"Tabela encontrada: {table_xpath}")
                
                # Tentar extrair dados usando JavaScript como alternativa
                js_script = f"""
                var table = document.evaluate('{table_xpath}', document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;
                var data = [];
                if (table) {{
                    var rows = table.querySelectorAll('tr');
                    for (var i = 0; i < rows.length; i++) {{
                        var cells = rows[i].querySelectorAll('td, th');
                        var rowData = [];
                        for (var j = 0; j < cells.length; j++) {{
                            rowData.push(cells[j].textContent.trim());
                        }}
                        data.push(rowData);
                    }}
                }}
                return data;
                """
                js_data = driver.execute_script(js_script)
                # print(f"Dados extraídos via JavaScript: {js_data}")
                
                # Se JavaScript conseguiu extrair dados, usar esses dados
                if js_data and any(any(cell for cell in row if cell) for row in js_data):
                    # print("Usando dados extraídos via JavaScript")
                    # Procurar pela coluna "Projeção (%)" nos dados JavaScript
                    projecao_col_index = -1
                    for i, row in enumerate(js_data):
                        for j, cell in enumerate(row):
                            if "Projeção" in cell and "%" in cell:
                                projecao_col_index = j
                                # print(f"Coluna 'Projeção (%)' encontrada no índice: {j}")
                                break
                        if projecao_col_index >= 0:
                            break
                    
                    # Se não encontrou pelo cabeçalho, procurar por padrões de valores percentuais
                    if projecao_col_index == -1:
                        for i, row in enumerate(js_data):
                            for j, cell in enumerate(row):
                                # Procurar por valores que parecem ser percentuais ou numéricos
                                if cell and (cell == '-' or 
                                           (',' in cell and any(c.isdigit() for c in cell)) or 
                                           ('.' in cell and any(c.isdigit() for c in cell)) or
                                           (cell.replace('-', '').replace(',', '').replace('.', '').isdigit())):
                                    projecao_col_index = j
                                    # print(f"Coluna de projeção detectada no índice: {j} com valor: '{cell}'")
                                    break
                            if projecao_col_index >= 0:
                                break
                    
                    # Extrair dados da coluna identificada
                    data = []
                    if projecao_col_index >= 0:
                        for i, row in enumerate(js_data[1:], 1):  # Pular cabeçalho
                            if len(row) > projecao_col_index:
                                cell_text = row[projecao_col_index]
                                if cell_text == "" or cell_text == "—" or cell_text == "–":
                                    cell_text = "-"
                                data.append(cell_text)
                                # print(f"JS - Linha {i}, célula {projecao_col_index}: '{cell_text}'")
                    else:
                        # Fallback: coletar todos os valores que parecem ser numéricos/percentuais
                        for row in js_data:
                            for cell in row:
                                if cell and (cell == '-' or 
                                           (',' in cell and any(c.isdigit() for c in cell) and not any(c.isalpha() for c in cell)) or 
                                           ('.' in cell and any(c.isdigit() for c in cell) and not any(c.isalpha() for c in cell)) or
                                           (cell.replace('-', '').replace(',', '').replace('.', '').isdigit())):
                                    data.append(cell)
                    
                    print(f"Dados finais extraídos via JavaScript: {data}")
                    return data
                
                # Se JavaScript não funcionou, continuar com método Selenium
                print("JavaScript não retornou dados válidos, tentando método Selenium...")
                
                # Extrair todas as linhas da tabela
                rows = table.find_elements(By.TAG_NAME, "tr")
                # print(f"Número de linhas encontradas: {len(rows)}")
                
                if len(rows) == 0:
                    return []
                
                # Encontrar o índice da coluna "Projeção (%)"
                projecao_col_index = -1
                header_row = rows[0]
                
                # Tentar encontrar nos cabeçalhos (th)
                headers = header_row.find_elements(By.TAG_NAME, "th")
                if headers:
                    # print(f"Cabeçalhos encontrados: {[h.text.strip() for h in headers]}")
                    for i, header in enumerate(headers):
                        header_text = header.text.strip()
                        if "Projeção" in header_text and "%" in header_text:
                            projecao_col_index = i
                            # print(f"Coluna 'Projeção (%)' encontrada no índice: {i}")
                            break
                
                # Se não encontrou nos th, tentar nos td da primeira linha
                if projecao_col_index == -1:
                    cells = header_row.find_elements(By.TAG_NAME, "td")
                    if cells:
                        # print(f"Células da primeira linha: {[c.text.strip() for c in cells]}")
                        for i, cell in enumerate(cells):
                            cell_text = cell.text.strip()
                            if "Projeção" in cell_text and "%" in cell_text:
                                projecao_col_index = i
                                # print(f"Coluna 'Projeção (%)' encontrada no índice: {i}")
                                break
                
                # Se ainda não encontrou, assumir uma posição baseada no padrão comum
                if projecao_col_index == -1:
                    # print("Coluna 'Projeção (%)' não encontrada nos cabeçalhos. Tentando detectar por conteúdo...")
                    # Vamos examinar todas as colunas de todas as linhas para encontrar dados que parecem percentuais
                    for row_idx, row in enumerate(rows):
                        cells = row.find_elements(By.TAG_NAME, "td")
                        if not cells:
                            cells = row.find_elements(By.TAG_NAME, "th")
                        
                        # Tentar múltiplos métodos para extrair texto
                        cell_texts = []
                        for cell in cells:
                            # Método 1: text
                            text1 = cell.text.strip()
                            # Método 2: get_attribute('textContent')
                            text2 = cell.get_attribute('textContent').strip() if cell.get_attribute('textContent') else ""
                            # Método 3: get_attribute('innerText')
                            text3 = cell.get_attribute('innerText').strip() if cell.get_attribute('innerText') else ""
                            # Método 4: innerHTML e limpar tags
                            innerHTML = cell.get_attribute('innerHTML')
                            import re
                            text4 = re.sub(r'<[^>]+>', '', innerHTML).strip() if innerHTML else ""
                            
                            # Usar o primeiro método que retornar conteúdo não vazio
                            final_text = text1 or text2 or text3 or text4
                            cell_texts.append(final_text)
                        
                        # print(f"Linha {row_idx}: {cell_texts}")
                        
                        for col_idx, cell_text in enumerate(cell_texts):
                            # Procurar por valores que parecem ser numéricos/percentuais
                            if cell_text and (cell_text == '-' or 
                                            (',' in cell_text and any(c.isdigit() for c in cell_text) and not any(c.isalpha() for c in cell_text)) or 
                                            ('.' in cell_text and any(c.isdigit() for c in cell_text) and not any(c.isalpha() for c in cell_text)) or
                                            (cell_text.replace('-', '').replace(',', '').replace('.', '').isdigit())):
                                if projecao_col_index == -1:
                                    projecao_col_index = col_idx
                                    # print(f"Possível coluna de projeção detectada no índice: {col_idx} com valor: '{cell_text}'")
                
                # Extrair dados da coluna identificada
                data = []
                
                if projecao_col_index >= 0:
                    # print(f"Extraindo dados da coluna índice: {projecao_col_index}")
                    for row_idx, row in enumerate(rows[1:], 1):  # Pular possível cabeçalho
                        cells = row.find_elements(By.TAG_NAME, "td")
                        if len(cells) > projecao_col_index:
                            cell = cells[projecao_col_index]
                            # Usar múltiplos métodos para extrair texto
                            text1 = cell.text.strip()
                            text2 = cell.get_attribute('textContent').strip() if cell.get_attribute('textContent') else ""
                            text3 = cell.get_attribute('innerText').strip() if cell.get_attribute('innerText') else ""
                            innerHTML = cell.get_attribute('innerHTML')
                            text4 = re.sub(r'<[^>]+>', '', innerHTML).strip() if innerHTML else ""
                            
                            cell_text = text1 or text2 or text3 or text4
                            if cell_text == "" or cell_text == "—" or cell_text == "–":
                                cell_text = "-"
                            # print(f"Linha {row_idx}, célula {projecao_col_index}: '{cell_text}'")
                            data.append(cell_text)
                else:
                    print("Não foi possível identificar a coluna de projeção. Coletando todos os valores que parecem percentuais...")
                    # Como fallback, coletar todos os valores que parecem ser percentuais
                    for row in rows:
                        cells = row.find_elements(By.TAG_NAME, "td")
                        for cell in cells:
                            # Usar múltiplos métodos para extrair texto
                            text1 = cell.text.strip()
                            text2 = cell.get_attribute('textContent').strip() if cell.get_attribute('textContent') else ""
                            text3 = cell.get_attribute('innerText').strip() if cell.get_attribute('innerText') else ""
                            innerHTML = cell.get_attribute('innerHTML')
                            text4 = re.sub(r'<[^>]+>', '', innerHTML).strip() if innerHTML else ""
                            
                            cell_text = text1 or text2 or text3 or text4
                            if cell_text and (cell_text == '-' or 
                                            (',' in cell_text and any(c.isdigit() for c in cell_text) and not any(c.isalpha() for c in cell_text)) or 
                                            ('.' in cell_text and any(c.isdigit() for c in cell_text) and not any(c.isalpha() for c in cell_text)) or
                                            (cell_text.replace('-', '').replace(',', '').replace('.', '').isdigit())):
                                if cell_text == "" or cell_text == "—" or cell_text == "–":
                                    cell_text = "-"
                                data.append(cell_text)
                
                print(f"Dados extraídos: {data}")
                return data
                
            except Exception as e:
                print(f"Erro ao extrair dados da tabela {table_xpath}: {e}")
                import traceback
                traceback.print_exc()
                return []
        
        # Extrair dados das duas tabelas
        tabela1_data = extract_projecao_column(table1_xpath)
        tabela2_data = extract_projecao_column(table2_xpath)
        
        # Função auxiliar para extrair cabeçalhos th da primeira tabela
        def extract_table_headers(table_xpath):
            try:
                table = driver.find_element(By.XPATH, table_xpath)
                headers = table.find_elements(By.TAG_NAME, "th")
                header_texts = []
                for header in headers:
                    # Usar múltiplos métodos para extrair texto
                    text1 = header.text.strip()
                    text2 = header.get_attribute('textContent').strip() if header.get_attribute('textContent') else ""
                    text3 = header.get_attribute('innerText').strip() if header.get_attribute('innerText') else ""
                    innerHTML = header.get_attribute('innerHTML')
                    text4 = re.sub(r'<[^>]+>', '', innerHTML).strip() if innerHTML else ""
                    
                    final_text = text1 or text2 or text3 or text4
                    if final_text:
                        header_texts.append(final_text)
                
                # print(f"Cabeçalhos da primeira tabela: {header_texts}")
                return header_texts
            except Exception as e:
                print(f"Erro ao extrair cabeçalhos: {e}")
                return []
        
        # Extrair cabeçalhos da primeira tabela
        table1_headers = extract_table_headers(table1_xpath)
        
        # Obter o mês atual em formato extenso
        # Importar as variáveis necessárias do escopo global
        global mes, m0_extenso
        mes_atual = m0_extenso[int(mes)]
        # print(f"Mês atual: {mes_atual}")
        
        # Verificar se o mês atual está contido em algum cabeçalho da primeira tabela
        mes_encontrado_tabela1 = False
        for header in table1_headers:
            if mes_atual in header.upper():
                mes_encontrado_tabela1 = True
                # print(f"Mês '{mes_atual}' encontrado no cabeçalho: '{header}'")
                break
        
        # Lógica condicional para retornar dados
        if mes_encontrado_tabela1:
            print(f"Retornando dados da primeira tabela (mês {mes_atual} encontrado)")
            return tabela1_data
        else:
            print(f"Retornando dados da segunda tabela (mês {mes_atual} não encontrado na primeira tabela)")
            return tabela2_data
        
    except TimeoutException:
        print("Timeout: Página não carregou completamente")
        return []
    except Exception as e:
        print(f"Erro durante o web scraping: {e}")
        return []
    finally:
        if driver:
            driver.quit()

# Função que mantém a interface original mas usa Selenium
def projecoes():
    """
    Função que mantém a mesma interface da função original,
    mas usa Selenium internamente para extrair os dados.
    """
    return web_scraping_anbima()


def ipca1ou2():

    if projecoes()[2] == '-':
        resultado =  'Usar IPCA 1'
        return float(projecoes()[int(resultado[-1])].replace(",", "."))

    else: 
        resultado =  'Usar IPCA 2'
        return float(projecoes()[int(resultado[-1])].replace(",", "."))





# definindo quais serão os meses que precisaremos de previsão do IPCA

meses_previsao = ipca.loc[ipca.inf_mensal.isna(), 'mes'].unique()
# realização da query + importação para dataframe
query = (f"SELECT distinct * FROM db_asset.dbo.ANBIMA_PROJECAO_IGP where tipo like '%IPCA%' and projecao is not null order by data")
proj_ipca = pd.read_sql(sql=query, con=conn)


# mudando as datas para string
proj_ipca['data'] = pd.to_datetime(proj_ipca['data'])
proj_ipca['data'] = proj_ipca['data'].dt.strftime('%Y-%m-%d')

# Adicionar valor hardcoded para fevereiro 2026 = 0.44%
if len(meses_previsao) > 0:
    mes_ultimo = meses_previsao[-1]
    # Extrair apenas o mês se for formato "YYYY-MM"
    if '-' in str(mes_ultimo):
        mes_ultimo_num = str(mes_ultimo).split('-')[-1]
    else:
        mes_ultimo_num = str(mes_ultimo).zfill(2)
    
    # Se for fevereiro (02), usar o IPCA efetivo de janeiro
    if mes_ultimo_num == '02':
        # Buscar IPCA efetivo de janeiro no ipca_base
        ipca_efetivo_janeiro = ipca_base.loc[ipca_base['data'].str.contains('-01-'), 'inf_mensal'].values
        
        if len(ipca_efetivo_janeiro) > 0:
            valor_fevereiro = ipca_efetivo_janeiro[-1]
            print(f"Usando IPCA efetivo de Janeiro 2026 para Fevereiro = {valor_fevereiro}%")
        else:
            valor_fevereiro = 0.33
            print(f"IPCA efetivo de janeiro não encontrado. Usando projeção = {valor_fevereiro}%")
        
        linha_add = pd.DataFrame({'data':[dmenos1] ,'mes_coleta' : [f'{m0_extenso[int(mes_ultimo_num)]} de {ano}'] ,'projecao': [valor_fevereiro],'validade': [dmenos1] ,'tipo': ['linha adicionada']})
        proj_ipca = pd.concat([proj_ipca, linha_add], ignore_index=True)
    elif len(proj_ipca.loc[proj_ipca['data'].str.contains(f'-{mes_ultimo_num}-')]) == 0:
        # Se não for fevereiro e também não existir no banco, busca no site
        linha_add = pd.DataFrame({'data':[dmenos1] ,'mes_coleta' : [f'{m0_extenso[int(mes_ultimo_num)]} de {ano}'] ,'projecao': [ipca1ou2()],'validade': [dmenos1] ,'tipo': ['linha adicionada']})
        proj_ipca = pd.concat([proj_ipca, linha_add], ignore_index=True)
    
#pegando as projeções para os meses de "meses_previsao" 
taxas = []
for month in meses_previsao:
    # Extrair apenas o mês se for formato "YYYY-MM"
    if '-' in str(month):
        mes_num = str(month).split('-')[-1]
    else:
        mes_num = str(month).zfill(2)
    
    valores = proj_ipca.loc[proj_ipca['data'].str.contains(f'-{mes_num}-'), 'projecao'].values
    if len(valores) > 0:
        taxas.append(valores[-1])
    else:
        # Se não encontrar para o mês, traz o último efetivo disponível
        ultimo_disponivel = proj_ipca['projecao'].iloc[-1]
        print(f"Aviso: Nenhum dado encontrado para o mês {mes_num}. Usando última projeção disponível: {ultimo_disponivel}")
        taxas.append(ultimo_disponivel)

# criando um dicionário com a taxa correspondente para cada mês
mes_taxa = {}
for m,t in zip(meses_previsao, taxas):

    mes_taxa[m] = t

# alimentando o df ipca com as taxas projedatas de IPCA
for m in meses_previsao:

    ipca.loc[ipca['mes']==m, 'inf_mensal'] = mes_taxa[m] 


# selecionando somente as colunas desejadas
ipca = ipca[['Dias Uteis', 'inf_mensal']]

#trocando nome da coluna data
ipca.rename(columns={'Dias Uteis': 'Data'}, inplace=True)

# mudando as taxas para decimal
ipca['IPCA'] = ipca['inf_mensal'] / 100

# selecionando somente as colunas desejadas
ipca = ipca[['Data','IPCA']]


################################## CDI##################################

# realização da query + importação para dataframe
query = ('SELECT * FROM db_asset.dbo.CDI_CETIP order by data')
cdi_base = pd.read_sql(sql=query, con=conn)

# mudando as datas para string
cdi_base['data'] = pd.to_datetime(cdi_base['data'])
cdi_base['data'] = cdi_base['data'].dt.strftime('%Y-%m-%d')

# deixando as colunas no padrão do código
cdi = cdi_base.rename(columns={'data': 'Data', 'valor':'CDI'})

# concertando o descasamento de 1 dia entre o SQL e a CapitâniaMailer, pois aqui consideraremos que o CapitâniaMailer é que está certo 
cdi['CDI'] = cdi['CDI'].shift(1)

# Seguindo padrão do CapitâniaMailer, onde temos CDI a partir de '2011-02-09'
cdi = cdi.loc[cdi['Data']>='2011-02-09']


################################## IPCA+i ##################################

# inserir a taxa pré do benchcmark Exemplos: 5.5  ;  7


def df_ipca(i):
    # índice acumulado

    i = i/100

    ipca_mais = ipca.copy()

    ipca_mais['Acum'] = np.cumprod(((1 + ipca_mais['IPCA'])**12*(1+i))**(1/252))

    return ipca_mais
# CDI + i


################################## CDI + i ##################################

def df_cdi(i):
    # índice acumulado

    i = i/100

    cdi_mais = cdi.copy()

    cdi_mais['Acum'] = (((1 + cdi_mais['CDI']/100)*(1 + i))**(1/252)).cumprod()

    return cdi_mais

cdi_acum = df_cdi(0)[['Data', 'Acum']]


################################## fundos q já tiveram mais de 1 benchmark ##################################

def troca_bench(fundo_):
    
    # descobrindo qual é o benchmark e qual é o valor da taxa pré-fixada

    if fundo_bench[fundo_][0][:4] == 'cdi+':

        taxa = float(fundo_bench[fundo_][0][fundo_bench[fundo_][0].find('+') + 1 : ])
        bench = cdi.copy()

    elif fundo_bench[fundo_][0][:5] == 'ipca+':

        taxa = float(fundo_bench[fundo_][0][fundo_bench[fundo_][0].find('+') + 1 : ])
        bench = ipca.copy()
    
    # filtrando a partir de data inicial do fundo
    bench = bench.loc[bench['Data']>=cota_inicial[fundo_][0]]
    # criando uma coluna que recebe o valor da taxa pré (a inicial)
    bench['taxa'] = taxa

    
    # troca da taxa pré-fixada a partir da data de mudança
    for d, i in zip(fundo_bench[fundo_][1::2], fundo_bench[fundo_][2::2]):
    
        bench.loc[bench['Data']>= d, 'taxa'] = i

    # calculo CDI(IPCA)+ 
    if fundo_bench[fundo_][0][:4] == 'cdi+':

        bench['Acum'] = np.cumprod(((1 + bench['CDI']/100)*(1 + bench['taxa']/100))**(1/252)) 

    elif fundo_bench[fundo_][0][:5] == 'ipca+':

        bench['Acum'] = np.cumprod(((1 + bench['IPCA']) ** 12 * (1 + bench['taxa']/100)) ** (1 / 252)) 
    
    return bench

################################## IFIX ##################################

query = (f'''
SELECT * FROM db_asset.dbo.IFIX order by data
''')

ifix = pd.read_sql(sql=query, con=conn)

ifix.columns = ['Data', 'Acum']

ifix['Data'] = pd.to_datetime(ifix['Data'])

ifix['Data'] = ifix['Data'].dt.strftime('%Y-%m-%d')

################################## Retorno em D.U. ##################################

# retorno do índice para um número dado de dias úteis

def bench_du(df_bench, num_dus):

    try:
        return df_bench.iloc[-1, -1] / df_bench.iloc[-1 - num_dus, -1] - 1
    except:
        return np.nan

################################## Retorno entre datas ##################################

def bench_delta(fundo_, df_bench, data_a, data_b):

    if data_a < cota_inicial[fundo_][0]:

        data_a = cota_inicial[fundo_][0]

    return df_bench.loc[df_bench['Data']==data_b, 'Acum'].values[0] / df_bench.loc[df_bench['Data']==data_a, 'Acum'].values[0] - 1



################################## Cotas Cap ##################################

query = (f'''
SELECT [Data], [Fundo], [Cota], [PL] FROM db_asset.dbo.COTAS_CAP order by data desc
''')

cotas_cap_base = pd.read_sql(sql=query, con=conn)

cotas_cap_base['Data'] = pd.to_datetime(cotas_cap_base['Data'])

cotas_cap_base['Data'] = cotas_cap_base['Data'].dt.strftime('%Y-%m-%d')

# separando esse DataFrame para usar na função de check PL
fundo_pl = cotas_cap_base[['Fundo','PL']]

# comtinuando
cotas_cap_base = cotas_cap_base[['Data', 'Fundo','Cota']]


################################## Cota Ajustada ##################################

cotas_cap_base.drop(cotas_cap_base[cotas_cap_base['Fundo'].str.contains('|'.join(fundos_com_ajuste))].index, inplace=True)

query = (f'''
SELECT [Data], [Fundo], [Cota_Ajustada] FROM db_asset.dbo.Cotas_Ret_Ajus order by data desc
''')

cotas_ajustadas = pd.read_sql(sql=query, con=conn)

cotas_ajustadas['Data'] = pd.to_datetime(cotas_ajustadas['Data'])

cotas_ajustadas['Data'] = cotas_ajustadas['Data'].dt.strftime('%Y-%m-%d')

cotas_ajustadas.rename(columns={'Cota_Ajustada': 'Cota'}, inplace=True)

conn.close()



################################## Base de Cotas Completa ##################################
cotas_cap = pd.concat([cotas_cap_base, cotas_ajustadas])


################################## Dados COTAS CAP por fundo ##################################

def cota_base(fundo_):

    dados_mailer = cotas_cap[cotas_cap['Fundo']== fundo_]
    dados_fundo = {}

    lista_resultados = []

    for d in [dmenos1,dmenos(2)]:

        resultado = dados_mailer.loc[dados_mailer['Data']==d, 'Cota'].empty

        if resultado:
            print(f'Cota do {fundo_} ref. {d} não está no COTAS CAP')
        lista_resultados.append(resultado)

    if any(lista_resultados):
        
        dados_fundo = {chave: np.nan for chave in ['valor_cota','var_dia', 'var_mes', 'var_ano', 'dia_cdi', 'mes_cdi', 'ano_cdi','p_cdi', 'p_cdi_mes', 'p_cdi_ano']} 
        
        return dados_fundo

    else:
            
        if fundo_ in fundos_mellon or fundo_ in fundos_itau or fundo_ in fundos_bradesco:

            #Cota
            dados_fundo['valor_cota'] =  dados_mailer.loc[dados_mailer['Data']==f'{ano}-{mes}-{dia}', 'Cota'].values[0]
            
            #Dia
            dados_fundo['var_dia'] =  dados_mailer['Cota'].values[0] / dados_mailer['Cota'].values[1] - 1
            
            # % do CDI
            dados_fundo['p_cdi'] =  dados_fundo['var_dia'] / (cdi_acum.iloc[-1,1] / cdi_acum.iloc[-2,1] - 1)
            
            #Mês
            try:
                dados_fundo['var_mes'] =  dados_fundo['valor_cota'] / dados_mailer.loc[dados_mailer['Data']==fechamento, 'Cota'].values[0] - 1
            except:
                dados_fundo['var_mes'] = dados_fundo['valor_cota'] / cota_inicial[fundo_][1] - 1

            #Ano
            try:
                dados_fundo['var_ano'] = dados_fundo['valor_cota'] / dados_mailer.loc[dados_mailer['Data']==ano_velho, 'Cota'].values[0] - 1
            except:
                dados_fundo['var_ano'] = dados_fundo['valor_cota'] / cota_inicial[fundo_][1] - 1
            
            
            # adicionado o cdi ao dicionário

            dados_fundo['dia_cdi'] = bench_du(cdi_acum, 1)

            if datetime.strptime(fechamento, '%Y-%m-%d') < datetime.strptime(cota_inicial[fundo_][0], '%Y-%m-%d'):
                dados_fundo['mes_cdi'] = bench_delta(fundo_, cdi_acum, cota_inicial[fundo_][0], dmenos1)
            else:
                dados_fundo['mes_cdi'] = bench_delta(fundo_, cdi_acum, fechamento, dmenos1)

            if datetime.strptime(ano_velho, '%Y-%m-%d') < datetime.strptime(cota_inicial[fundo_][0], '%Y-%m-%d'):
                dados_fundo['ano_cdi'] = bench_delta(fundo_, cdi_acum, cota_inicial[fundo_][0], dmenos1)
            else:
                dados_fundo['ano_cdi'] = bench_delta(fundo_, cdi_acum, ano_velho, dmenos1)

            
            return dados_fundo
            

        elif fundo_ in fundos_btg:
            
            #Cota
            dados_fundo['valor_cota'] =  dados_mailer.loc[dados_mailer['Data']==f'{ano}-{mes}-{dia}', 'Cota'].values[0]
            
            #Dia
            dados_fundo['var_dia'] =  dados_mailer['Cota'].values[0] / dados_mailer['Cota'].values[1] - 1
            # % do CDI dia
            dados_fundo['p_cdi'] =  dados_fundo['var_dia'] / (cdi_acum.iloc[-1,1] / cdi_acum.iloc[-2,1] - 1)
            
            #Mês
            try:
                dados_fundo['var_mes'] =  dados_fundo['valor_cota'] / dados_mailer.loc[dados_mailer['Data']==fechamento, 'Cota'].values[0] - 1
            except:
                dados_fundo['var_mes'] = dados_fundo['valor_cota'] / cota_inicial[fundo_][1] - 1
            # % do CDI mes
            dados_fundo['p_cdi_mes'] =  dados_fundo['var_mes'] / bench_delta(fundo_, cdi_acum, fechamento, dmenos1)

            #Ano
            try:
                dados_fundo['var_ano'] = dados_mailer.loc[dados_mailer['Data']==f'{ano}-{mes}-{dia}', 'Cota'].values[0] / dados_mailer.loc[dados_mailer['Data']==ano_velho, 'Cota'].values[0] - 1
            except:
                dados_fundo['var_ano'] = dados_mailer.loc[dados_mailer['Data']==f'{ano}-{mes}-{dia}', 'Cota'].values[0] / cota_inicial[fundo_][1] -1
            # % do CDI ano
            dados_fundo['p_cdi_ano'] = dados_fundo['var_ano'] / bench_delta(fundo_, cdi_acum, ano_velho, dmenos1)
        
            return dados_fundo
        
        
        elif fundo_ in fundos_xp:
            
            #Cota
            dados_fundo['valor_cota'] =  dados_mailer.loc[dados_mailer['Data']==f'{ano}-{mes}-{dia}', 'Cota'].values[0]
            
            #Dia
            
            dados_fundo['var_dia'] =  dados_mailer['Cota'].values[0] / dados_mailer['Cota'].values[1] - 1

            #Mês
            try:
                dados_fundo['var_mes'] =  dados_fundo['valor_cota'] / dados_mailer.loc[dados_mailer['Data']==fechamento, 'Cota'].values[0] - 1
            except:
                dados_fundo['var_mes'] = dados_fundo['valor_cota'] / cota_inicial[fundo_][1] - 1
            
            #Ano
            try:
                dados_fundo['var_ano'] = dados_mailer.loc[dados_mailer['Data']==f'{ano}-{mes}-{dia}', 'Cota'].values[0] / dados_mailer.loc[dados_mailer['Data']==ano_velho, 'Cota'].values[0] - 1
            except:
                dados_fundo['var_ano'] = dados_mailer.loc[dados_mailer['Data']==f'{ano}-{mes}-{dia}', 'Cota'].values[0] / cota_inicial[fundo_][1] -1
            
            return dados_fundo


################################## Dados PL ##################################
# retorna df com PL hoje e PL médio respectivamente

def pl(fundo_):

    dados_mailer = fundo_pl.loc[fundo_pl['Fundo']== fundo_, 'PL']

    pl_hoje =  dados_mailer.values[0]

    pl_medio = dados_mailer.iloc[0:252].mean()

    return pd.DataFrame([pl_hoje , pl_medio])


################################## Retorno em D.U. ##################################
# cálculo dos retornos para o fundo e número de Dias Úteis solicitado ->>> sem necessidade de batimento

def fundo_du(fundo_, num_dus):

    retornos = cotas_cap.loc[cotas_cap['Fundo']==fundo_ , ['Data','Fundo','Cota']]

    try:
        if (retornos.iloc[0, 2] != 0) and (retornos.iloc[num_dus, 2] != 0):
            return retornos.iloc[0, 2] / retornos.iloc[num_dus, 2] - 1
        else:
            return np.nan
    except:
        return np.nan

################################## Retorno entre datas ##################################
def fundo_delta(fundo_, data_a, data_b):

    df = cotas_cap.iloc[:, :3]
    df_filtro = df.loc[df['Fundo']==fundo_]

    if df_filtro.loc[df_filtro['Data']==data_a, 'Cota'].empty:

        return np.nan 

    else:

        cota_b = df_filtro.loc[df_filtro['Data']==data_b, 'Cota'].values[0]
        cota_a = df_filtro.loc[df_filtro['Data']==data_a, 'Cota'].values[0]

        if (cota_b != 0) and (cota_a != 0):
            return cota_b / cota_a - 1
        else:
            return np.nan




################################## Retorno no Mês para o fundo e benchmark ##################################
def mtd(fundo_, bench_):

    if datetime.strptime(fechamento, '%Y-%m-%d') < datetime.strptime(cota_inicial[fundo_][0], '%Y-%m-%d'):

        
        mtd_fundo = fundo_delta(fundo_, cota_inicial[fundo_][0], dmenos1)
        mtd_bench = bench_delta(fundo_, bench_ , cota_inicial[fundo_][0], dmenos1)

        return ( mtd_fundo, mtd_bench )
    
    else:

        mtd_fundo = fundo_delta(fundo_, fechamento, dmenos1)
        mtd_bench = bench_delta(fundo_, bench_ , fechamento, dmenos1)

        return ( mtd_fundo, mtd_bench )


################################## Retorno YTD para o fundo e benchmark ##################################
def ytd(fundo_, bench_):

    if datetime.strptime(ano_velho, '%Y-%m-%d') < datetime.strptime(cota_inicial[fundo_][0], '%Y-%m-%d'):

        
        ytd_fundo = fundo_delta(fundo_, cota_inicial[fundo_][0], dmenos1)
        ytd_bench = bench_delta(fundo_, bench_ , cota_inicial[fundo_][0], dmenos1)

        return ( ytd_fundo, ytd_bench )
    
    else:

        ytd_fundo = fundo_delta(fundo_, ano_velho, dmenos1)
        ytd_bench = bench_delta(fundo_, bench_ , ano_velho, dmenos1)

        return ( ytd_fundo, ytd_bench )

################################## Retorno dos últimos dois anos bench ##################################
# Retornará uma tupla com o retorno do ano passado e retrasado respectivamente. Se o fundo for mais novo, retorna NaN no lugar da rentabilidade
def ret_anos_bench(fundo_, df_bench):

    ano_inicio = int(cota_inicial[fundo_][0][:4])
    ano_limite = int(dmenos1[:4])-2

    
    if ano_inicio < ano_limite:

        try:        
            bench_ano_velho = df_bench.loc[df_bench['Data']==ano_velho, df_bench.columns[-1]].values[0]
            bench_end_ano_menos2 = df_bench.loc[df_bench['Data']==end_ano_menos2, df_bench.columns[-1]].values[0]
            bench_end_ano_menos3 = df_bench.loc[df_bench['Data']==end_ano_menos3, df_bench.columns[-1]].values[0]
            
            return (bench_ano_velho / bench_end_ano_menos2 - 1, bench_end_ano_menos2 / bench_end_ano_menos3 - 1)
        except:
            return (np.nan , np.nan)

    elif ano_inicio == ano_limite:

        try:
            bench_ano_velho = df_bench.loc[df_bench['Data']==ano_velho, df_bench.columns[-1]].values[0]
            bench_end_ano_menos2 = df_bench.loc[df_bench['Data']==end_ano_menos2, df_bench.columns[-1]].values[0]
            bench_inicio = df_bench.loc[df_bench['Data']==cota_inicial[fundo_][0], df_bench.columns[-1]].values[0]

            return (bench_ano_velho / bench_end_ano_menos2 - 1, bench_end_ano_menos2 / bench_inicio - 1)
        except:
            return (np.nan , np.nan)

    elif ano_inicio - ano_limite == 1:

        try:
            bench_ano_velho = df_bench.loc[df_bench['Data']==ano_velho, df_bench.columns[-1]].values[0]
            bench_inicio = df_bench.loc[df_bench['Data']==cota_inicial[fundo_][0], df_bench.columns[-1]].values[0]
            
            return (bench_ano_velho / bench_inicio - 1, np.nan)
        
        except:

            return (np.nan , np.nan)
    

    else: # Então o fundo se iniciou no ano atual

        return (np.nan , np.nan)
# Retornará uma tupla com o retorno do ano passado e retrasado respectivamente. Se o fundo for mais novo, retorna NaN no lugar da rentabilidade


################################## Retorno dos últimos dois anos fundo ##################################
def ret_anos(fundo_):

    ano_inicio = int(cota_inicial[fundo_][0][:4])
    ano_limite = int(dmenos1[:4])-2

    cotas = cotas_cap.loc[cotas_cap['Fundo']== fundo_, ['Data','Fundo','Cota']]


    if ano_inicio < ano_limite:

        try:        
            cota_ano_velho = cotas.loc[cotas['Data']==ano_velho, 'Cota'].values[0]
            cota_end_ano_menos2 = cotas.loc[cotas['Data']==end_ano_menos2, 'Cota'].values[0]
            cota_end_ano_menos3 = cotas.loc[cotas['Data']==end_ano_menos3, 'Cota'].values[0]
            
            return (cota_ano_velho / cota_end_ano_menos2 - 1, cota_end_ano_menos2 / cota_end_ano_menos3 - 1)
        except:
            return (np.nan , np.nan)

    elif ano_inicio == ano_limite:

        try:
            cota_ano_velho = cotas.loc[cotas['Data']==ano_velho, 'Cota'].values[0]
            cota_end_ano_menos2 = cotas.loc[cotas['Data']==end_ano_menos2, 'Cota'].values[0]
            cota_inicio = cota_inicial[fundo_][1]

            return (cota_ano_velho / cota_end_ano_menos2 - 1, cota_end_ano_menos2 / cota_inicio - 1)
        except:
            return (np.nan , np.nan)

    elif ano_inicio - ano_limite == 1:

        try:
            cota_ano_velho = cotas.loc[cotas['Data']==ano_velho, 'Cota'].values[0]
            cota_inicio = cota_inicial[fundo_][1]
            
            return (cota_ano_velho / cota_inicio - 1, np.nan)
        
        except:

            return (np.nan , np.nan)
    

    else: # Então o fundo se iniciou no ano atual

        return (np.nan , np.nan)
    
## rentabilidade dos dias anteriores
################################## Ajustes do DF final ##################################

## funções para formatação dos dados que aparecerão no PDF
pos_neg = lambda x: x if x>=0 else x

ajuste_bench = lambda x: -1*x if x<0 else x

def mes_ano(data):
    data_formatada = datetime.strptime(data, '%Y-%m-%d')
    data_formatada = data_formatada.strftime('%b-%y')
    return data_formatada

################################## fórmulas usadas para o Batimento ##################################

# poderá ser usado para descobrir quais dos floats tem menos casas decimais, e utilizar o número de casas decimais para o ROUND da função batimento
def n_casasdecimais(num_float):

    string_test = str(num_float)

    return len(string_test[string_test.find('.') + 1:])

# verifica se obtivemos o valor no PL e do PL Médio
def bat_pl(fundo):

    if pl(fundo).isna().any().values[0]:

        return False
    
    else:

        return True


# verifica se a cota e as rentabilidades do fundo batem ao confrontar os dados da Mailer e CapitaniMailer com a Carteira do Fundo 
def batimento(fundo_):
        
    if fundo_ in fundos_com_ajuste:
        return True
    
    else:

        lista = []

        fundo_carteira = cota_carteira(fundo_)

        fundo_base = cota_base(fundo_)

        for k in fundo_carteira.keys():

            num = min( n_casasdecimais(fundo_base[k]) , n_casasdecimais(fundo_carteira[k]))

            if k == 'valor_cota':

                num = min( num, 8)
            
            elif k in ['var_dia', 'var_mes', 'var_ano', 'dia_cdi', 'mes_cdi', 'ano_cdi']:

                num  = min( num, 4)
            
            elif k in ['p_cdi', 'p_cdi_mes', 'p_cdi_ano']:

                num  = min( num, 2)

            lista.append( round(float(fundo_base[k]), num) == round(fundo_carteira[k], num) )

        if all(lista):

            return True
        
        else:
            
            lista = []

            fundo_carteira = {k : remove_trailing_zeros(format(cota_carteira(fundo_)[k], ".20f")) for k in cota_carteira(fundo_)}

            fundo_base = {k : remove_trailing_zeros(format(cota_base(fundo_)[k], ".20f")) for k in cota_base(fundo_)} 

            for k in fundo_carteira.keys():

                num = min( n_casasdecimais(fundo_base[k]) , n_casasdecimais(fundo_carteira[k]))

                if k == 'valor_cota':

                    num = min( num, 8)
                
                elif k in ['var_dia', 'var_mes', 'var_ano', 'dia_cdi', 'mes_cdi', 'ano_cdi']:

                    num  = min( num, 4)
                
                elif k in ['p_cdi', 'p_cdi_mes', 'p_cdi_ano']:

                    num  = min( num, 2)
                
                lista.append( str(fundo_base[k]) [ : str(fundo_base[k]).find('.') + 1 + num ] == str(fundo_carteira[k]) [ : str(fundo_carteira[k]).find('.') + 1 + num ])
            
            if all(lista):
                return True
            else:
                return False


# procura a cota do fundo para uma data específica
def get_cota(fundo_, data):
    return cotas_cap.loc[(cotas_cap['Fundo']==fundo_) & (cotas_cap['Data']==data), 'Cota'].values[0]


#retorna o benchmark do fundo em formato dataframe
def retorna_bench(fundo_):

    #definição do bench mark
    if fundo_bench[fundo_] == 'cdi':

        bench_ = cdi_acum.copy()
    
    elif fundo_bench[fundo_][:4] == 'cdi+':

        taxa = float(fundo_bench[fundo_][fundo_bench[fundo_].find('+') + 1 : ])
        bench_ = df_cdi(taxa)
    
    elif fundo_bench[fundo_] == 'ifix':

        bench_ = ifix.copy()
    
    elif fundo_bench[fundo_][:5] == 'ipca+':

        taxa = float(fundo_bench[fundo_][fundo_bench[fundo_].find('+') + 1 : ])
        bench_ = df_ipca(taxa)
    
    elif fundo_bench[fundo_] == 'imab5':
        
        bench_ = imab5_acum.copy()

    elif fundo_bench[fundo_] == 'imab':
        
        bench_ = imab_acum.copy()

    elif fundo_ in mais_de_1_bench:

        bench_ = troca_bench(fundo_)

    return bench_


################## CHECK BENCHMARK ##################

def check_bench(fundo_):

    lista_condicoes = []
    
    bench = retorna_bench(fundo_)

    # primeira linha é da data D-1 ?
    if bench.iloc[-1,0] == dmenos1:
        lista_condicoes.append('ok')
    else:
        lista_condicoes.append('erro')

    # o valor tá OK?
    ultima_linha = bench.iloc[-1,1:]
    na_ou_zero = (ultima_linha == 0).any() or ultima_linha.isna().any()

    if na_ou_zero:
        lista_condicoes.append('erro')
    else:
        lista_condicoes.append('ok')
    
    return 'erro' in lista_condicoes




############### Check Cotas ##################

def check_cotas(fundo_):
    
    df_cotas_fundo = cotas_cap.loc[cotas_cap['Fundo']==fundo_ , ['Data','Fundo','Cota']]
    lista = []
    mensagem = None
    if  cota_inicial[fundo_][0] > fechamento:
        dias_a_verificar =    [dmenos(1),dmenos(2),dmenos(3),dmenos(4),dmenos(5)]
    elif cota_inicial[fundo_][0] > ano_velho:
        dias_a_verificar =    [dmenos(1),dmenos(2),dmenos(3),dmenos(4),dmenos(5), fechamento]
    else:
        dias_a_verificar = [dmenos(1),dmenos(2),dmenos(3),dmenos(4),dmenos(5), fechamento, ano_velho]

    for data in dias_a_verificar:
        resultado = True
        if data in df_cotas_fundo.iloc[:,0].values:

            valor_cota =  df_cotas_fundo.loc[df_cotas_fundo['Data']==data,'Cota']

            if valor_cota.values[0]  == 0 or valor_cota.isna().values[0]:
                mensagem = f'Cota de {data} igual a 0 ou NaN no COTAS_CAP para o fundo {fundo_}'
                resultado =  False
            
        else:
            mensagem = f'Data {data} não consta no COTAS_CAP para o fundo {fundo_}'
            resultado =  False
        lista.append((data, resultado))
        
    return [all([dr[1] for dr in lista]) , mensagem]


################## Gerador DF ##################

def gerador_df(fundo):
    
    bench = retorna_bench(fundo)

    #troca cota inicial dos fundos com amortização
    if fundo in fundos_com_ajuste:
        cota_inicial[fundo][1] = cotas_cap.loc[(cotas_cap['Fundo']==fundo)& (cotas_cap['Data']==cota_inicial[fundo][0])].values[0][-1]
    else:
        pass

    #definição do modelo 
    # modelo 1
    if fundo in modelo_1:
        
        dados = [
            [dmenos1,cota_base(fundo)['valor_cota'], fundo_du(fundo, 1), bench_du(bench, 1),pos_neg(fundo_du(fundo, 1) / bench_du(bench, 1))],
            [dmenos(2), get_cota(fundo, dmenos(2)),fundo_delta(fundo, dmenos(3), dmenos(2)), bench_delta(fundo, bench,dmenos(3), dmenos(2)), pos_neg(fundo_delta(fundo, dmenos(3), dmenos(2)) / bench_delta(fundo, bench,dmenos(3), dmenos(2)))],
            [dmenos(3), get_cota(fundo, dmenos(3)),fundo_delta(fundo, dmenos(4), dmenos(3)), bench_delta(fundo, bench,dmenos(4), dmenos(3)), pos_neg(fundo_delta(fundo, dmenos(4), dmenos(3)) / bench_delta(fundo, bench,dmenos(4), dmenos(3)))],
            [dmenos(4), get_cota(fundo, dmenos(4)),fundo_delta(fundo, dmenos(5), dmenos(4)), bench_delta(fundo, bench,dmenos(5), dmenos(4)), pos_neg(fundo_delta(fundo, dmenos(5), dmenos(4)) / bench_delta(fundo, bench,dmenos(5), dmenos(4)))],
            [dmenos(5), get_cota(fundo, dmenos(5)),fundo_delta(fundo, dmenos(6), dmenos(5)), bench_delta(fundo, bench,dmenos(6), dmenos(5)), pos_neg(fundo_delta(fundo, dmenos(6), dmenos(5)) / bench_delta(fundo, bench,dmenos(6), dmenos(5)))],    
            [mes_ano(dmenos1) , '',mtd(fundo, bench)[0], mtd(fundo, bench)[1], mtd(fundo, bench)[0]/ mtd(fundo, bench)[1]],
            ['Últimos 30 dias', '',fundo_du(fundo, 21), bench_du(bench, 21), (fundo_du(fundo, 21)) / (bench_du(bench, 21))],
            ['Últimos 90 dias', '',fundo_du(fundo, 63), bench_du(bench, 63), (fundo_du(fundo, 63)) / (bench_du(bench, 63))],
            ['Últimos 180 dias', '',fundo_du(fundo, 126), bench_du(bench, 126), (fundo_du(fundo, 126)) / (bench_du(bench, 126))],
            ['Últimos 360 dias', '',fundo_du(fundo, 252), bench_du(bench, 252), (fundo_du(fundo, 252)) / (bench_du(bench, 252))],
            [f'Ano {int(ano)}', '',ytd(fundo, bench)[0], ytd(fundo, bench)[1], ytd(fundo, bench)[0] / ytd(fundo, bench)[1]],
            [f'Ano {int(ano) - 1}', '',ret_anos(fundo)[0], ret_anos_bench(fundo, bench)[0], ret_anos(fundo)[0] / ret_anos_bench(fundo, bench)[0]],
            [f'Ano {int(ano) - 2}', '',ret_anos(fundo)[1], ret_anos_bench(fundo, bench)[1], ret_anos(fundo)[1] / ret_anos_bench(fundo, bench)[1]],
            ['Acumulado²', '',cota_base(fundo)['valor_cota'] / cota_inicial[fundo][1] -1, bench_delta(fundo, bench,cota_inicial[fundo][0], dmenos1), (cota_base(fundo)['valor_cota'] / cota_inicial[fundo][1] -1) / bench_delta(fundo, bench,cota_inicial[fundo][0], dmenos1)]
        ]

    # modelo 2
    elif fundo in modelo_2:

        dados = [
            [dmenos1,cota_base(fundo)['valor_cota'], fundo_du(fundo, 1), bench_du(bench, 1), bench_du(cdi_acum, 1),pos_neg(fundo_du(fundo, 1) / bench_du(cdi_acum, 1))],
            [dmenos(2), get_cota(fundo, dmenos(2)),fundo_delta(fundo, dmenos(3), dmenos(2)), bench_delta(fundo, bench,dmenos(3), dmenos(2)), bench_delta(fundo, cdi_acum,dmenos(3), dmenos(2)),pos_neg(fundo_delta(fundo, dmenos(3), dmenos(2)) / bench_delta(fundo, cdi_acum,dmenos(3), dmenos(2)))],
            [dmenos(3), get_cota(fundo, dmenos(3)),fundo_delta(fundo, dmenos(4), dmenos(3)), bench_delta(fundo, bench,dmenos(4), dmenos(3)), bench_delta(fundo, cdi_acum,dmenos(4), dmenos(3)),pos_neg(fundo_delta(fundo, dmenos(4), dmenos(3)) / bench_delta(fundo, cdi_acum,dmenos(4), dmenos(3)))],
            [dmenos(4), get_cota(fundo, dmenos(4)),fundo_delta(fundo, dmenos(5), dmenos(4)), bench_delta(fundo, bench,dmenos(5), dmenos(4)), bench_delta(fundo, cdi_acum,dmenos(5), dmenos(4)),pos_neg(fundo_delta(fundo, dmenos(5), dmenos(4)) / bench_delta(fundo, cdi_acum,dmenos(5), dmenos(4)))],
            [dmenos(5), get_cota(fundo, dmenos(5)),fundo_delta(fundo, dmenos(6), dmenos(5)), bench_delta(fundo, bench,dmenos(6), dmenos(5)), bench_delta(fundo, cdi_acum,dmenos(6), dmenos(5)),pos_neg(fundo_delta(fundo, dmenos(6), dmenos(5)) / bench_delta(fundo, cdi_acum,dmenos(6), dmenos(5)))],    
            [mes_ano(dmenos1), '', fundo_delta(fundo,fechamento,dmenos1), bench_delta(fundo, bench,fechamento,dmenos1), bench_delta(fundo, cdi_acum,fechamento,dmenos1),fundo_delta(fundo,fechamento,dmenos1)/ bench_delta(fundo, cdi_acum,fechamento,dmenos1)],
            ['Últimos 30 dias', '', fundo_du(fundo, 21), bench_du(bench, 21), bench_du(cdi_acum, 21),(fundo_du(fundo, 21)) / (bench_du(cdi_acum, 21))],
            ['Últimos 90 dias', '', fundo_du(fundo, 63), bench_du(bench, 63), bench_du(cdi_acum, 63), (fundo_du(fundo, 63)) / (bench_du(cdi_acum, 63))],
            ['Últimos 180 dias', '', fundo_du(fundo, 126), bench_du(bench, 126), bench_du(cdi_acum, 126), (fundo_du(fundo, 126)) / (bench_du(cdi_acum, 126))],
            ['Últimos 360 dias', '', fundo_du(fundo, 252), bench_du(bench, 252), bench_du(cdi_acum, 252),(fundo_du(fundo, 252)) / (bench_du(cdi_acum, 252))],
            [f'Ano {int(ano)}', '', ytd(fundo, bench)[0], ytd(fundo, bench)[1], ytd(fundo, cdi_acum)[1] ,ytd(fundo, bench)[0] / ytd(fundo, cdi_acum)[1]],
            [f'Ano {int(ano) - 1}', '',ret_anos(fundo)[0], ret_anos_bench(fundo, bench)[0], ret_anos_bench(fundo, cdi_acum)[0], ret_anos(fundo)[0] / ret_anos_bench(fundo, cdi_acum)[0]],
            [f'Ano {int(ano) - 2}', '',ret_anos(fundo)[1], ret_anos_bench(fundo, bench)[1], ret_anos_bench(fundo, cdi_acum)[1], ret_anos(fundo)[1] / ret_anos_bench(fundo, cdi_acum)[1]],
            ['Acumulado²', '',cota_base(fundo)['valor_cota'] / cota_inicial[fundo][1] -1, bench_delta(fundo, bench,cota_inicial[fundo][0], dmenos1), bench_delta(fundo, cdi_acum,cota_inicial[fundo][0], dmenos1),(cota_base(fundo)['valor_cota'] / cota_inicial[fundo][1] -1) / bench_delta(fundo, cdi_acum,cota_inicial[fundo][0], dmenos1)]
        ]


    # modelo 3
    elif fundo in modelo_3:

        dados = [
            [dmenos1,cota_base(fundo)['valor_cota'], fundo_du(fundo, 1), bench_du(ifix, 1)],
            [dmenos(2), get_cota(fundo, dmenos(2)),fundo_delta(fundo, dmenos(3), dmenos(2)), bench_delta(fundo,ifix,dmenos(3), dmenos(2))],
            [dmenos(3), get_cota(fundo, dmenos(3)),fundo_delta(fundo, dmenos(4), dmenos(3)), bench_delta(fundo,ifix,dmenos(4), dmenos(3))],
            [dmenos(4), get_cota(fundo, dmenos(4)),fundo_delta(fundo, dmenos(5), dmenos(4)), bench_delta(fundo,ifix,dmenos(5), dmenos(4))],
            [dmenos(5), get_cota(fundo, dmenos(5)),fundo_delta(fundo, dmenos(6), dmenos(5)), bench_delta(fundo,ifix,dmenos(6), dmenos(5))],    
            [mes_ano(dmenos1), '',fundo_delta(fundo,fechamento,dmenos1), bench_delta(fundo,ifix,fechamento,dmenos1)],
            ['Últimos 30 dias', '',fundo_du(fundo, 21), bench_du(ifix, 21)],
            ['Últimos 90 dias', '',fundo_du(fundo, 63), bench_du(ifix, 63)],
            ['Últimos 180 dias', '',fundo_du(fundo, 126), bench_du(ifix, 126)],
            ['Últimos 360 dias', '',fundo_du(fundo, 252), bench_du(ifix, 252)],
            [f'Ano {int(ano)}', '',ytd(fundo, bench)[0], ytd(fundo, ifix)[1]],
            [f'Ano {int(ano) - 1}', '',ret_anos(fundo)[0], ret_anos_bench(fundo, ifix)[0]],
            [f'Ano {int(ano) - 2}', '',ret_anos(fundo)[1], ret_anos_bench(fundo, ifix)[1]],
            ['Acumulado²', '',cota_base(fundo)['valor_cota'] / cota_inicial[fundo][1] -1, bench_delta(fundo,ifix,cota_inicial[fundo][0], dmenos1)]
        ]    


    # modelo 4
    elif fundo in modelo_4:

        dados = [
            [dmenos1,cota_base(fundo)['valor_cota'], fundo_du(fundo, 1), bench_du(cdi_acum, 1), bench_du(bench, 1), pos_neg(fundo_du(fundo, 1) / bench_du(cdi_acum, 1)), pos_neg(fundo_du(fundo, 1) / bench_du(bench, 1))],
            [dmenos(2), get_cota(fundo, dmenos(2)),fundo_delta(fundo, dmenos(3), dmenos(2)), bench_delta(fundo,cdi_acum,dmenos(3), dmenos(2)), bench_delta(fundo, bench,dmenos(3), dmenos(2)),pos_neg(fundo_delta(fundo, dmenos(3), dmenos(2)) / bench_delta(fundo,cdi_acum,dmenos(3), dmenos(2))),pos_neg(fundo_delta(fundo, dmenos(3), dmenos(2)) / bench_delta(fundo, bench,dmenos(3), dmenos(2)))],
            [dmenos(3), get_cota(fundo, dmenos(3)),fundo_delta(fundo, dmenos(4), dmenos(3)), bench_delta(fundo,cdi_acum,dmenos(4), dmenos(3)), bench_delta(fundo, bench,dmenos(4), dmenos(3)),pos_neg(fundo_delta(fundo, dmenos(4), dmenos(3)) / bench_delta(fundo,cdi_acum,dmenos(4), dmenos(3))),pos_neg(fundo_delta(fundo, dmenos(4), dmenos(3)) / bench_delta(fundo, bench,dmenos(4), dmenos(3)))],
            [dmenos(4), get_cota(fundo, dmenos(4)),fundo_delta(fundo, dmenos(5), dmenos(4)), bench_delta(fundo,cdi_acum,dmenos(5), dmenos(4)), bench_delta(fundo, bench,dmenos(5), dmenos(4)),pos_neg(fundo_delta(fundo, dmenos(5), dmenos(4)) / bench_delta(fundo,cdi_acum,dmenos(5), dmenos(4))),pos_neg(fundo_delta(fundo, dmenos(5), dmenos(4)) / bench_delta(fundo, bench,dmenos(5), dmenos(4)))],
            [dmenos(5), get_cota(fundo, dmenos(5)),fundo_delta(fundo, dmenos(6), dmenos(5)), bench_delta(fundo,cdi_acum,dmenos(6), dmenos(5)), bench_delta(fundo, bench,dmenos(6), dmenos(5)),pos_neg(fundo_delta(fundo, dmenos(6), dmenos(5)) / bench_delta(fundo,cdi_acum,dmenos(6), dmenos(5))),pos_neg(fundo_delta(fundo, dmenos(6), dmenos(5)) / bench_delta(fundo, bench,dmenos(6), dmenos(5)))],    
            [mes_ano(dmenos1), '', mtd(fundo, bench)[0], mtd(fundo, cdi_acum)[1], mtd(fundo, bench)[1],mtd(fundo, cdi_acum)[0]/ mtd(fundo, cdi_acum)[1] ,mtd(fundo, bench)[0]/ mtd(fundo, bench)[1]],
            ['Últimos 30 dias', '',fundo_du(fundo, 21), bench_du(cdi_acum, 21), bench_du(bench, 21), (fundo_du(fundo, 21)) / (bench_du(cdi_acum, 21)), (fundo_du(fundo, 21)) / (bench_du(bench, 21))],
            ['Últimos 90 dias', '',fundo_du(fundo, 63), bench_du(cdi_acum, 63), bench_du(bench, 63), (fundo_du(fundo, 63)) / (bench_du(cdi_acum, 63)), (fundo_du(fundo, 63)) / (bench_du(bench, 63))],
            ['Últimos 180 dias' , '',fundo_du(fundo, 126), bench_du(cdi_acum, 126), bench_du(bench, 126), (fundo_du(fundo, 126)) / (bench_du(cdi_acum, 126)), (fundo_du(fundo, 126)) / (bench_du(bench, 126))],
            ['Últimos 360 dias', '',fundo_du(fundo, 252), bench_du(cdi_acum, 252), bench_du(bench, 252), (fundo_du(fundo, 252)) / (bench_du(cdi_acum, 252)), (fundo_du(fundo, 252)) / (bench_du(bench, 252))],
            [f'Ano {int(ano)}', '',ytd(fundo, bench)[0], ytd(fundo, cdi_acum)[1], ytd(fundo, bench)[1] ,ytd(fundo, bench)[0] / ytd(fundo, cdi_acum)[1] ,ytd(fundo, bench)[0] / ytd(fundo, bench)[1]],
            [f'Ano {int(ano) - 1}', '',ret_anos(fundo)[0], ret_anos_bench(fundo, cdi_acum)[0], ret_anos_bench(fundo, bench)[0], ret_anos(fundo)[0] / ret_anos_bench(fundo, cdi_acum)[0], ret_anos(fundo)[0] / ret_anos_bench(fundo, bench)[0]],
            [f'Ano {int(ano) - 2}', '',ret_anos(fundo)[1], ret_anos_bench(fundo, cdi_acum)[1], ret_anos_bench(fundo, bench)[1], ret_anos(fundo)[1] / ret_anos_bench(fundo, cdi_acum)[1], ret_anos(fundo)[1] / ret_anos_bench(fundo, bench)[1]],
            ['Acumulado²', '',cota_base(fundo)['valor_cota'] / cota_inicial[fundo][1] -1, bench_delta(fundo,cdi_acum,cota_inicial[fundo][0], dmenos1), bench_delta(fundo, bench,cota_inicial[fundo][0], dmenos1),(cota_base(fundo)['valor_cota'] / cota_inicial[fundo][1] -1) / bench_delta(fundo,cdi_acum,cota_inicial[fundo][0], dmenos1),(cota_base(fundo)['valor_cota'] / cota_inicial[fundo][1] -1) / bench_delta(fundo, bench,cota_inicial[fundo][0], dmenos1)]
        ]

    elif fundo in modelo_5:

        dados = [
            [dmenos1,cota_base(fundo)['valor_cota'], fundo_du(fundo, 1), bench_du(cdi_acum, 1), bench_du(imab_acum, 1)],
            [dmenos(2), get_cota(fundo, dmenos(2)),fundo_delta(fundo, dmenos(3), dmenos(2)), bench_delta(fundo,cdi_acum,dmenos(3), dmenos(2)), bench_delta(fundo,imab_acum,dmenos(3), dmenos(2))],
            [dmenos(3), get_cota(fundo, dmenos(3)),fundo_delta(fundo, dmenos(4), dmenos(3)), bench_delta(fundo,cdi_acum,dmenos(4), dmenos(3)), bench_delta(fundo,imab_acum,dmenos(4), dmenos(3))],
            [dmenos(4), get_cota(fundo, dmenos(4)),fundo_delta(fundo, dmenos(5), dmenos(4)), bench_delta(fundo,cdi_acum,dmenos(5), dmenos(4)), bench_delta(fundo,imab_acum,dmenos(5), dmenos(4))],
            [dmenos(5), get_cota(fundo, dmenos(5)),fundo_delta(fundo, dmenos(6), dmenos(5)), bench_delta(fundo,cdi_acum,dmenos(6), dmenos(5)), bench_delta(fundo,imab_acum,dmenos(6), dmenos(5))],    
            [mes_ano(dmenos1) , '',mtd(fundo, bench)[0], mtd(fundo, cdi_acum)[1], mtd(fundo, bench)[1]],
            ['Últimos 30 dias', '',fundo_du(fundo, 21), bench_du(cdi_acum, 21), bench_du(imab_acum, 21)],
            ['Últimos 90 dias', '',fundo_du(fundo, 63), bench_du(cdi_acum, 63), bench_du(imab_acum, 63)],
            ['Últimos 180 dias', '',fundo_du(fundo, 126), bench_du(cdi_acum, 126), bench_du(imab_acum, 126)],
            ['Últimos 360 dias', '',fundo_du(fundo, 252), bench_du(cdi_acum, 252), bench_du(imab_acum, 252)],
            [f'Ano {int(ano)}', '',ytd(fundo, cdi_acum)[0], ytd(fundo, cdi_acum)[1], ytd(fundo, imab_acum)[1]],
            [f'Ano {int(ano) - 1}', '',ret_anos(fundo)[0], ret_anos_bench(fundo, cdi_acum)[0], ret_anos_bench(fundo, imab_acum)[0]],
            [f'Ano {int(ano) - 2}', '',ret_anos(fundo)[1], ret_anos_bench(fundo, cdi_acum)[1], ret_anos_bench(fundo, imab_acum)[1]],
            ['Acumulado²', '',cota_base(fundo)['valor_cota'] / cota_inicial[fundo][1] -1, bench_delta(fundo,cdi_acum,cota_inicial[fundo][0], dmenos1), bench_delta(fundo,imab_acum,cota_inicial[fundo][0], dmenos1)]
        ]


    elif fundo in modelo_fapes:

        dados = [
            [dmenos1,cota_base(fundo)['valor_cota'], fundo_du(fundo, 1), bench_du(imab_acum, 1), fundo_du(fundo, 1) - bench_du(imab_acum, 1), bench_du(imab5_acum, 1), fundo_du(fundo, 1) - bench_du(imab5_acum, 1)],
            [dmenos(2),get_cota(fundo, dmenos(2)), fundo_delta(fundo, dmenos(3), dmenos(2)), bench_delta(fundo,imab_acum,dmenos(3), dmenos(2)), fundo_delta(fundo, dmenos(3), dmenos(2))- bench_delta(fundo,imab_acum,dmenos(3), dmenos(2)),bench_delta(fundo,imab5_acum,dmenos(3), dmenos(2)), fundo_delta(fundo, dmenos(3), dmenos(2))- bench_delta(fundo,imab5_acum,dmenos(3), dmenos(2))],
            [dmenos(3),get_cota(fundo, dmenos(3)), fundo_delta(fundo, dmenos(4), dmenos(3)), bench_delta(fundo,imab_acum,dmenos(4), dmenos(3)), fundo_delta(fundo, dmenos(4), dmenos(3))- bench_delta(fundo,imab_acum,dmenos(4), dmenos(3)),bench_delta(fundo,imab5_acum,dmenos(4), dmenos(3)), fundo_delta(fundo, dmenos(4), dmenos(3))- bench_delta(fundo,imab5_acum,dmenos(4), dmenos(3))],
            [dmenos(4),get_cota(fundo, dmenos(4)), fundo_delta(fundo, dmenos(5), dmenos(4)), bench_delta(fundo,imab_acum,dmenos(5), dmenos(4)), fundo_delta(fundo, dmenos(5), dmenos(4))- bench_delta(fundo,imab_acum,dmenos(5), dmenos(4)),bench_delta(fundo,imab5_acum,dmenos(5), dmenos(4)), fundo_delta(fundo, dmenos(5), dmenos(4))- bench_delta(fundo,imab5_acum,dmenos(5), dmenos(4))],
            [dmenos(5),get_cota(fundo, dmenos(5)), fundo_delta(fundo, dmenos(6), dmenos(5)), bench_delta(fundo,imab_acum,dmenos(6), dmenos(5)), fundo_delta(fundo, dmenos(6), dmenos(5))- bench_delta(fundo,imab_acum,dmenos(6), dmenos(5)),bench_delta(fundo,imab5_acum,dmenos(6), dmenos(5)), fundo_delta(fundo, dmenos(6), dmenos(5))- bench_delta(fundo,imab5_acum,dmenos(6), dmenos(5))],
            [mes_ano(dmenos1) , '',fundo_delta(fundo,fechamento,dmenos1), bench_delta(fundo,imab_acum,fechamento,dmenos1),fundo_delta(fundo,fechamento,dmenos1)- bench_delta(fundo,imab_acum,fechamento,dmenos1), bench_delta(fundo,imab5_acum,fechamento,dmenos1), fundo_delta(fundo,fechamento,dmenos1)- bench_delta(fundo,imab5_acum,fechamento,dmenos1)],
            ['Últimos 30 dias', '',fundo_du(fundo, 21), bench_du(imab_acum, 21),fundo_du(fundo, 21)- bench_du(imab_acum, 21),bench_du(imab5_acum, 21),fundo_du(fundo, 21)-  bench_du(imab5_acum, 21)],
            ['Últimos 90 dias', '',fundo_du(fundo, 63), bench_du(imab_acum, 63),fundo_du(fundo, 63)- bench_du(imab_acum, 63),bench_du(imab5_acum, 63),fundo_du(fundo, 63)-  bench_du(imab5_acum, 63)],
            ['Últimos 180 dias', '',fundo_du(fundo, 126), bench_du(imab_acum, 126),fundo_du(fundo, 126)- bench_du(imab_acum, 126),bench_du(imab5_acum, 126),fundo_du(fundo, 126)-  bench_du(imab5_acum, 126)],
            ['Últimos 360 dias', '',fundo_du(fundo, 252), bench_du(imab_acum, 252),fundo_du(fundo, 252)- bench_du(imab_acum, 252),bench_du(imab5_acum, 252),fundo_du(fundo, 252)-  bench_du(imab5_acum, 252)],
            [f'Ano {int(ano)}', '',ytd(fundo, bench)[0], ytd(fundo, imab_acum)[1],ytd(fundo, imab_acum)[0] - ytd(fundo, imab_acum)[1], ytd(fundo, imab5_acum)[1], ytd(fundo, imab5_acum)[0] - ytd(fundo, imab5_acum)[1]],
            [f'Ano {int(ano) - 1}', '',ret_anos(fundo)[0], ret_anos_bench(fundo, imab_acum)[0],ret_anos(fundo)[0]- ret_anos_bench(fundo, imab_acum)[0], ret_anos_bench(fundo, imab5_acum)[0],ret_anos(fundo)[0]- ret_anos_bench(fundo, imab5_acum)[0]],
            [f'Ano {int(ano) - 2}', '',ret_anos(fundo)[1], ret_anos_bench(fundo, imab_acum)[1],ret_anos(fundo)[1]- ret_anos_bench(fundo, imab_acum)[1], ret_anos_bench(fundo, imab5_acum)[1],ret_anos(fundo)[1]- ret_anos_bench(fundo, imab5_acum)[1]],
            ['Acumulado²', '',cota_base(fundo)['valor_cota'] / cota_inicial[fundo][1] -1, bench_delta(fundo,imab_acum,cota_inicial[fundo][0], dmenos1),cota_base(fundo)['valor_cota'] / cota_inicial[fundo][1] -1 - bench_delta(fundo,imab_acum,cota_inicial[fundo][0], dmenos1), bench_delta(fundo,imab5_acum,cota_inicial[fundo][0], dmenos1),cota_base(fundo)['valor_cota'] / cota_inicial[fundo][1] -1 - bench_delta(fundo,imab5_acum,cota_inicial[fundo][0], dmenos1)]
        ]


    # modelo FUNCEF
    elif fundo in modelo_funcef:

        dados = [
            [dmenos1,cota_base(fundo)['valor_cota'], fundo_du(fundo, 1), bench_du(bench, 1),pos_neg(fundo_du(fundo, 1) / bench_du(bench, 1)), bench_du(ifix, 1)],
            [dmenos(2), get_cota(fundo, dmenos(2)),fundo_delta(fundo, dmenos(3), dmenos(2)), bench_delta(fundo, bench,dmenos(3), dmenos(2)), pos_neg(fundo_delta(fundo, dmenos(3), dmenos(2)) / bench_delta(fundo, bench,dmenos(3), dmenos(2))), bench_delta(fundo,ifix,dmenos(3), dmenos(2))],
            [dmenos(3), get_cota(fundo, dmenos(3)),fundo_delta(fundo, dmenos(4), dmenos(3)), bench_delta(fundo, bench,dmenos(4), dmenos(3)), pos_neg(fundo_delta(fundo, dmenos(4), dmenos(3)) / bench_delta(fundo, bench,dmenos(4), dmenos(3))), bench_delta(fundo,ifix,dmenos(4), dmenos(3))],
            [dmenos(4), get_cota(fundo, dmenos(4)),fundo_delta(fundo, dmenos(5), dmenos(4)), bench_delta(fundo, bench,dmenos(5), dmenos(4)), pos_neg(fundo_delta(fundo, dmenos(5), dmenos(4)) / bench_delta(fundo, bench,dmenos(5), dmenos(4))), bench_delta(fundo,ifix,dmenos(5), dmenos(4))],
            [dmenos(5), get_cota(fundo, dmenos(5)),fundo_delta(fundo, dmenos(6), dmenos(5)), bench_delta(fundo, bench,dmenos(6), dmenos(5)), pos_neg(fundo_delta(fundo, dmenos(6), dmenos(5)) / bench_delta(fundo, bench,dmenos(6), dmenos(5))), bench_delta(fundo,ifix,dmenos(6), dmenos(5))],    
            [mes_ano(dmenos1) , '',fundo_delta(fundo,fechamento,dmenos1), bench_delta(fundo, bench,fechamento,dmenos1), fundo_delta(fundo,fechamento,dmenos1)/ bench_delta(fundo, bench,fechamento,dmenos1), bench_delta(fundo,ifix,fechamento,dmenos1)],
            ['Últimos 30 dias', '',fundo_du(fundo, 21), bench_du(bench, 21), (fundo_du(fundo, 21)) / (bench_du(bench, 21)), bench_du(ifix, 21)],
            ['Últimos 90 dias', '',fundo_du(fundo, 63), bench_du(bench, 63), (fundo_du(fundo, 63)) / (bench_du(bench, 63)), bench_du(ifix, 63)],
            ['Últimos 180 dias', '',fundo_du(fundo, 126), bench_du(bench, 126), (fundo_du(fundo, 126)) / (bench_du(bench, 126)), bench_du(ifix, 126)],
            ['Últimos 360 dias', '',fundo_du(fundo, 252), bench_du(bench, 252), (fundo_du(fundo, 252)) / (bench_du(bench, 252)),fundo_du(fundo, 252)],
            [f'Ano {int(ano)}', '',ytd(fundo, bench)[0], ytd(fundo, bench)[1], ytd(fundo, bench)[0] / ytd(fundo, bench)[1], ytd(fundo, ifix)[1]],
            [f'Ano {int(ano) - 1}', '',ret_anos(fundo)[0], ret_anos_bench(fundo, bench)[0], ret_anos(fundo)[0] / ret_anos_bench(fundo, bench)[0], ret_anos_bench(fundo, ifix)[0]],
            [f'Ano {int(ano) - 2}', '',ret_anos(fundo)[1], ret_anos_bench(fundo, bench)[1], ret_anos(fundo)[1] / ret_anos_bench(fundo, bench)[1], ret_anos_bench(fundo, ifix)[1]],
            ['Acumulado²', '',cota_base(fundo)['valor_cota'] / cota_inicial[fundo][1] -1, bench_delta(fundo, bench,cota_inicial[fundo][0], dmenos1), (cota_base(fundo)['valor_cota'] / cota_inicial[fundo][1] -1) / bench_delta(fundo, bench,cota_inicial[fundo][0], dmenos1), bench_delta(fundo,ifix,cota_inicial[fundo][0], dmenos1)]
        ]


    cotas = cotas_cap.loc[(cotas_cap['Fundo'] == fundo) & (cotas_cap['Data']>=cota_inicial[fundo][0]), ['Data','Fundo','Cota'] ] 


    if len(cotas) < 22:

        del dados[6:10]

    elif len(cotas) < 64:

        del dados[7:10]

    elif len(cotas) < 127:

        del dados[8:10]

    elif len(cotas) < 253:

        del dados[9]


    if np.isnan(dados[4][2]) or np.isinf(dados[4][2]):
        del dados[4]

        if np.isnan(dados[3][2]) or np.isinf(dados[4][2]):
            del dados[4]

            if np.isnan(dados[2][2]) or np.isinf(dados[4][2]):
                del dados[4]

                if np.isnan(dados[1][2]) or np.isinf(dados[4][2]):
                    del dados[4]
                
                else:
                    pass              

            else:
                pass

        else:
            pass

    else:
        pass



    if np.isnan(dados[-2][2]) or np.isinf(dados[-2][2]):
        del dados[-2]
        
    else:
        pass

    if np.isnan(dados[-2][2]) or np.isinf(dados[-2][2]):
        del dados[-2]

    else:
        pass

    return pd.DataFrame(dados)


################################## verificação da rentabilidade de anos anteriore ##################################
# verifica se foi calculada a rentabilidade para os anos anteriores (se não foi, talvez faltem cotas de anos anteriores no COTAS_CAP)
def check_rent_anos(fundo_):

    #calculando qual DEVERIA ser o ano que consta na última linha do DataFrame

    if int(cota_inicial[fundo_][0][:4]) <= int(ano) - 2:
        ultima_linha = f'Ano {int(ano) - 2}'

    elif int(cota_inicial[fundo_][0][:4]) == int(ano) - 1:
        ultima_linha = f'Ano {int(ano) - 1}'

    elif int(cota_inicial[fundo_][0][:4]) == int(ano):
        ultima_linha = f'Ano {int(ano)}'

    # O ano esperado é o ano que consta na última linha?

    return ultima_linha == gerador_df(fundo_).iloc[-2,0]





################################## E-mail: informações e função de Envio ##################################

def get_email_infos():

    infos_email_ = {}
    infos_email_grupos = {}
    df_infos_email = pd.read_excel(r"X:\BDM\Novo Modelo de Carteiras\Tipo_Fundos.xlsx",
                                    sheet_name='destinatarios_mailer',
                                    engine='openpyxl',
                                    usecols='A:F')

    # --- PASSO 3: Identificar Colunas de Detalhes ---
    # Colunas base que não são detalhes
    fundo_grupo = ['fundo', 'grupo']

    # As colunas de detalhes são todas as outras colunas
    # É importante manter a ordem original delas
    parametros_email = [col for col in df_infos_email.columns if col not in fundo_grupo]

    # --- PASSO 4: Processar as Linhas do DataFrame ---
    for index, row in df_infos_email.iterrows():
        fundo = row['fundo']
        grupo = row['grupo']

        # Coletar os valores das colunas de detalhe para esta linha
        # Substituir NaN (Not a Number), que o pandas usa para células vazias, por strings vazias
        lista_parametros = [row[col] for col in parametros_email]
        
        # Converter explicitamente NaN para string vazia, pois df.fillna('') pode não ser suficiente dependendo da leitura
        infos_lista_parametros = []
        for item in lista_parametros:
            if pd.isna(item):
                infos_lista_parametros.append('')
            else:
                infos_lista_parametros.append(str(item) if not isinstance(item, str) else item) # Garante que é string

        # Reconstruir a estrutura do dicionário

        if grupo == 'unico':
            # Esta era uma entrada direta no dicionário (ex: 'BNY11279': [...])
            infos_email_[fundo] = infos_lista_parametros
        else:
            # Esta era uma entrada aninhada (ex: 'CAPITANIA CW1': {'bradesco': [...]})
            infos_email_grupos[grupo] = infos_lista_parametros
            infos_email_[fundo] = infos_email_grupos

    return infos_email_


# dicionário que contém as seguintes infos: [0]"NOME DO FUNDO" , [1]"TO", [2]"CC", [3]"BCC"
infos_email = get_email_infos()

# função de Envio
def send_outlook(fund):

    assinatura = """
    Atenciosamente,<br><br>

    <div style="font-size: 11px; font-family: Verdana;">
    <span style="color:#1C57A8;">Relações com Investidores</span><br>
    <span style="color:#1C57A8;"><b>Capitânia Investimentos </b></span><br>
    <span style="color:#1C57A8;">Tel: 55-11-2853-8888</span><br>
    <span style="color:#1C57A8;"">www.capitaniainvestimentos.com.br </span><br><br>
    <span style="color:#1C57A8;""><b>Informação confidencial para uso exclusivo pelo destinatário da mensagem. Confidential information for exclusive use by recipient.</b></span>
    </div>
    """

    fonte = "<body style='font-family:Calibri;font-size:12pt;'>"

    if fund == 'CAPITANIA CW1':

        outlook = win32.Dispatch('Outlook.Application')
        email = outlook.CreateItem(0)
        email.To = infos_email[fund]['bradesco'][1]
        email.CC = infos_email[fund]['bradesco'][2]
        email.BCC = infos_email[fund]['bradesco'][3]
        email.Subject = f"COTA DIÁRIA | {infos_email[fund]['bradesco'][0]}"
        email.HTMLBody = fonte + f"Prezados,<br><br>Segue anexo o relatório diário de rentabilidade do {infos_email[fund]['bradesco'][0]}.<br><br>" + assinatura
        email.Attachments.Add(f"{diretorio}\\PDFs\\{fund}_{ano}{mes}{dia}.pdf")
        email.Display()

        outlook = win32.Dispatch('Outlook.Application')
        email = outlook.CreateItem(0)
        email.To = infos_email[fund]['itau'][1]
        email.CC = infos_email[fund]['itau'][2]
        email.BCC = infos_email[fund]['itau'][3]
        email.Subject = f"COTA DIÁRIA | {infos_email[fund]['itau'][0]}"
        email.HTMLBody = fonte + f"Prezados,<br><br>Segue anexo o relatório diário de rentabilidade do {infos_email[fund]['itau'][0]}.<br><br>" + assinatura
        email.Attachments.Add(f"{diretorio}\\PDFs\\{fund}_{ano}{mes}{dia}.pdf")
        email.Display()

    else:
        outlook = win32.Dispatch('Outlook.Application')
        email = outlook.CreateItem(0)
        email.To = infos_email[fund][1]
        email.CC = infos_email[fund][2]
        email.BCC = infos_email[fund][3]
        email.Subject = f"COTA DIÁRIA | {infos_email[fund][0]}"
        email.HTMLBody = fonte + f"Prezados,<br><br>Segue anexo o relatório diário de rentabilidade do {infos_email[fund][0]}.<br><br>" + assinatura
        email.Attachments.Add(f"{diretorio}\\PDFs\\{fund}_{ano}{mes}{dia}.pdf")
        email.Display()


################################## PDF ##################################



def salvar_pdf(xlsx_path: str):
    out_dir = Path(f"{diretorio}\\PDFs")
    pdf = convert(
        source=xlsx_path,
        output_dir=out_dir,
        soft=0        
    )
    return pdf        

################################## MAILER ##################################

def mailer(fundos_selecionados):
    
    for f in fundos_selecionados:
        
        # Teste 0: Verifica se as cotas estão corretas.
        if not check_cotas(f)[0]:
            print(check_cotas(f)[1])
            continue  # Interrompe o loop imediatamente
        
        # Teste 1: Verifica o benchmark (só é executado se o teste0 tiver sido "ok")
        if check_bench(f):
            print(f'{f}: Tabela do {fundo_bench[f]} sem dados para o dia {dmenos1}')
            continue

        # Teste 3: Verifica o bat_pl.
        if not bat_pl(f):
            print(f +': PL não bateu')
            continue

        # Teste 4: Verifica se há algum NaN no dataframe gerado.
        # Aqui, a lógica é invertida: se houver algum NaN (ou seja, se o teste for True), então há erro.

        # Teste 2: Verifica o batimento.
        if not batimento(f):
            print(f +': Carteira não bate com COTAS_CAP')
            continue

        try:
            df = gerador_df(f)
        except:
            print(f + ': não foi possível gerar a tabela')
            continue
        
        if True in df.isna().any().values.tolist():
            print(f +': Existem valores NaN na tabela')
            continue
        
        
        out_dir   = Path(rf"{diretorio}\PDFs")
        out_dir.mkdir(parents=True, exist_ok=True)

        try:
            # Carregar o arquivo existente
            # Abre o arquivo Excel

            file_path = f"{diretorio}\\templates\\{f} - template.xlsx"
            
            # Carregar o workbook preservando formatação
            wb = openpyxl.load_workbook(file_path)
            plan = wb["Email"]

            #############################################################################
            # BLOCO 1: Inserir valores do DataFrame df.iloc[0:5]
            valores1 = df.iloc[0:5]
            # Procurar célula com "Data" no intervalo A10:C20
            for row in plan.iter_rows(min_row=10, max_row=20, min_col=1, max_col=3):
                for cell in row:
                    if cell.value == 'Data':
                        header_row = cell.row
                        col = cell.column
                        break
                else:
                    continue
                break
            # Preencher os valores abaixo do cabeçalho sem alterar a formatação
            for i, row_data in enumerate(valores1.values, start=header_row+1):
                for j, value in enumerate(row_data, start=col):
                    plan.cell(row=i, column=j, value=value)

            #############################################################################
            # BLOCO 2: Inserir valores do DataFrame df.iloc[5:]
            valores2 = df.iloc[5:]
            # Procurar célula com "Referência" no intervalo A17:C27
            for row in plan.iter_rows(min_row=17, max_row=27, min_col=1, max_col=3):
                for cell in row:
                    if cell.value == 'Referência':
                        header_row = cell.row
                        col = cell.column
                        break
                else:
                    continue
                break
            for i, row_data in enumerate(valores2.values, start=header_row+1):
                for j, value in enumerate(row_data, start=col):
                    plan.cell(row=i, column=j, value=value)

            #############################################################################
            # BLOCO 3: Inserir valores retornados pela função pl(f)
            valores3 = pl(f)
            # Procurar célula com "Patrimônio Líquido (R$)" no intervalo A27:C39
            for row in plan.iter_rows(min_row=27, max_row=39, min_col=1, max_col=3):
                for cell in row:
                    if cell.value == 'Patrimônio Líquido (R$)':
                        header_row = cell.row
                        col = cell.column
                        break
                else:
                    continue
                break
            for i, row_data in enumerate(valores3.values, start=header_row+1):
                for j, value in enumerate(row_data, start=col+3):
                    plan.cell(row=i, column=j, value=value)

            # Inserir demais valores sem alterar formatação
            plan.cell(row=header_row+4, column=col+1, value=dmenos1)
            plan.cell(row=header_row+5, column=col+1, value=cota_inicial[f][0])

            # Salvar o arquivo atualizado
            wb.save(file_path)

            pdf_path = convert(
                source=str(file_path),
                output_dir=str(out_dir),
                soft=0                 # 0 = MS Office; 1 = LibreOffice
            )

            # Renomear para {fundo}_{AAAAMMDD}.pdf se necessário
            desired_pdf = out_dir / f"{f}_{ano}{mes}{dia}.pdf"
            if Path(pdf_path).name != desired_pdf.name:
                Path(pdf_path).replace(desired_pdf)
                pdf_path = desired_pdf            

            #se já editou e fechou a planilha, é pq está pronto pra ser enviado

            send_outlook(f)
            
            print(f'{f}: Batimento OK. E-mail Enviado')

            # except:
            #     print(f"{f} - Não foi possível imprimir o pdf com o Libreoffice")
            
        except Exception as e:
            import traceback
            print(f'{f} - Problema: {e}')
            traceback.print_exc()



################################## Opções de input para o usuário ##################################
input_adm = {'1': fundos_mellon,
  '2': fundos_xp,
  '3': fundos_btg,
  '4': fundos_bradesco,
  '5': fundos_itau,
  't': tudo}


################################## Menu para escolher os fundos ##################################

class Menu:
    def __init__(self, master):
        self.master = master
        self.master.geometry("450x550")
        self.master.title("Fundos")

        self.options = input_adm[input('1-Mellon, 2-XP, 3-BTG, 4-Bradesco, 5-Itaú, t-tudo  ')] 
        self.selected_options = []

        self.create_widgets()

    def create_widgets(self):
        self.label = tk.Label(self.master, text="Selecione os fundos desejados e clique no botão para adicioná-los à lista de envio.")
        self.label.pack(pady=10)

        self.listbox = tk.Listbox(self.master, selectmode=tk.MULTIPLE)
        for option in self.options:
            self.listbox.insert(tk.END, option)
        self.listbox.pack(pady=10)

        self.button = tk.Button(self.master, text="Adicionar fundos", command=self.add_selected_options)
        self.button.pack(pady=10)

        self.selected_options_label = tk.Label(self.master, text="Fundos selecionados: ")
        self.selected_options_label.pack(pady=10)

        self.selected_options_listbox = tk.Listbox(self.master)
        self.selected_options_listbox.pack(pady=10)

    def add_selected_options(self):
        selected_indices = self.listbox.curselection()
        for index in selected_indices:
            option = self.listbox.get(index)
            self.selected_options.append(option)
        self.selected_options_listbox.delete(0, tk.END)
        for option in self.selected_options:
            self.selected_options_listbox.insert(tk.END, option)
        print(self.selected_options)

root = tk.Tk()
menu = Menu(root)
root.mainloop()

##################### MAILER #####################
mailer(menu.selected_options)



